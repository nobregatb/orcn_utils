from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os


# -------- Configurações --------
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
FILES_FOLDER = r"C:\Users\tbnobrega\OneDrive - ANATEL\Anatel\_ORCN"
PROFILE_DIR = os.path.abspath("./meu_perfil_chrome")
LOG_REQUERIMENTOS = os.path.join(FILES_FOLDER, 'ORCN - Documentação pertinente.xlsx')
LISTA_URL = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"

# -------- Funções --------
def criar_pasta_se_nao_existir(req):
    num, ano = req.split("/")
    requerimento = rf"Requerimentos\{ano}.{num}"
    full_path = os.path.join(FILES_FOLDER, requerimento)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
    return full_path

def atualizar_excel(rows):
    novosAnalises = []
    indicesNovos = []
    wb = load_workbook(LOG_REQUERIMENTOS)
    ws = wb['Requerimentos-Análise']
    tabRequerimentos = ws.tables['tabRequerimentos']

    linhas_existentes = [tuple(r) for r in ws.iter_rows(values_only=True)]

    for i, row in enumerate(rows, start=1):
        cols = row.query_selector_all("td") 
        dados = [col.inner_text().strip() for col in cols]
        dados[0] = 'AUTOMATICO'

        if dados[9] == 'Em Análise':
            novalinha = dados[:10]
            if tuple(novalinha) not in linhas_existentes:
                ws.append(novalinha)
                print(f"Linha adicionada: {novalinha}")
                novosAnalises.append(novalinha[1])
                indicesNovos.append(i)
                criar_pasta_se_nao_existir(novalinha[1])
            else:
                print("Linha já existe, não adicionada.")

    # Atualiza o range da tabela
    ultima_linha = ws.max_row
    col_inicio, linha_inicio = tabRequerimentos.ref.split(":")[0][0], tabRequerimentos.ref.split(":")[0][1:]
    col_fim = tabRequerimentos.ref.split(":")[1][0]
    tabRequerimentos.ref = f"{col_inicio}{linha_inicio}:{col_fim}{ultima_linha}"

    wb.save(LOG_REQUERIMENTOS)
    wb.close()
    return indicesNovos, novosAnalises

def salvarPDFs():
    # Aqui você coloca a lógica de salvar PDFs da aba aberta
    print("📄 Salvando PDFs... (função ainda precisa ser implementada)")

# -------- Script principal --------
with sync_playwright() as p:
    browser = p.chromium.launch_persistent_context(
        PROFILE_DIR,
        headless=False,
        executable_path=CHROME_PATH,
        args=["--start-maximized"]
    )
    page = browser.new_page()
    page.goto(LISTA_URL)
    page.wait_for_load_state("load")
    page.click("#menuForm\\:todos")
    page.wait_for_load_state("load")

    rows = page.query_selector_all("css=#form\\:tarefasTable_data tr")
    print(f"🔎 {len(rows)} processos na caixa de entrada")

    # Atualiza Excel e cria pastas
    novos_indices, novas_analises = atualizar_excel(rows)
    print("✅ Excel atualizado")

    # Processa cada linha nova
    for i, row in enumerate(rows, start=1):
        if i not in novos_indices:
            continue

        btn = row.query_selector("button[title='Visualizar em Tela cheia']")
        if not btn:
            print(f"⚠ Botão não encontrado na linha {i}")
            continue       

        btn.click(modifiers=["Control"])
        

        # Chama função de salvar PDFs
        salvarPDFs()

        # Fecha a aba nova e volta para a lista
        
        print("↩ Retornando para a lista de processos...")

    input("\nPressione ENTER para encerrar...")
    browser.close()
