from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os
import time
import re
import sys
from datetime import datetime

# Detecta se está sendo executado como executável PyInstaller
def is_bundled():
    return getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS')

# Detecta se foi chamado com parâmetro debug
#debug_mode = len(sys.argv) > 1 and sys.argv[1].lower() == 'debug'
debug_mode = 'debug'

# Configuração de caminhos
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

# Define FILES_FOLDER baseado no modo de execução
if debug_mode:
    # Modo debug: usa caminho fixo do desenvolvedor
    FILES_FOLDER = r"C:\Users\tbnobrega\OneDrive - ANATEL\Anatel\_ORCN"
    print("🐛 MODO DEBUG ATIVADO - Usando caminho de desenvolvimento")
else:
    if is_bundled():
        # Executável: usa diretório onde o .exe está localizado
        FILES_FOLDER = os.path.dirname(sys.executable)
        print(f"📦 MODO EXECUTÁVEL - Usando diretório: {FILES_FOLDER}")
    else:
        # Script Python normal: usa diretório do arquivo .py
        FILES_FOLDER = os.path.dirname(os.path.abspath(__file__))
        print(f"🐍 MODO SCRIPT - Usando diretório: {FILES_FOLDER}")

# Configura caminhos dependentes
PROFILE_DIR = os.path.join(FILES_FOLDER, "meu_perfil_chrome")
LOG_REQUERIMENTOS = os.path.join(FILES_FOLDER, 'ORCN.xlsx')

def req_para_fullpath(req):
    num, ano = req.split("/")
    requerimento = rf"Requerimentos\_{ano}.{num}"
    full_path = os.path.join(FILES_FOLDER, requerimento)
    return full_path
    
def criar_pasta_se_nao_existir(req):
    full_path = req_para_fullpath(req)
    if not os.path.exists(full_path):
        os.makedirs(full_path, exist_ok=True)
        print(f"   📁 Pasta criada: {full_path}")
    return full_path


def atualizar_excel(rows) -> list:
    """Atualiza planilha com novos requerimentos em análise (apenas no modo debug)"""
    if not debug_mode:
        # Modo não-debug: processa todos os requerimentos sem verificar Excel
        print("📋 Modo produção: processando todos os requerimentos da lista")
        todos_requerimentos = []
        for i, row in enumerate(rows, start=1):
            try:
                cols = row.query_selector_all("td") 
                if len(cols) >= 2:
                    requerimento = cols[1].inner_text().strip()
                    todos_requerimentos.append(requerimento)
                    print(f"   📋 Requerimento encontrado: {requerimento}")
            except Exception as e:
                print(f"   ⚠ Erro ao ler linha {i}: {str(e)[:50]}")
        
        print(f"\n📊 Total de requerimentos a processar: {len(todos_requerimentos)}")
        return todos_requerimentos
    
    # Modo debug: funcionalidade original com Excel
    print("🐛 Modo debug: verificando planilha Excel...")
    if not os.path.exists(LOG_REQUERIMENTOS):
        print(f"⚠ Planilha não encontrada: {LOG_REQUERIMENTOS}")
        print("📋 Processando todos os requerimentos da lista")
        todos_requerimentos = []
        for i, row in enumerate(rows, start=1):
            try:
                cols = row.query_selector_all("td") 
                if len(cols) >= 2:
                    requerimento = cols[1].inner_text().strip()
                    todos_requerimentos.append(requerimento)
            except Exception as e:
                print(f"   ⚠ Erro ao ler linha {i}: {str(e)[:50]}")
        return todos_requerimentos
    
    novosAnalises = []
    wb = load_workbook(LOG_REQUERIMENTOS)
    ws = wb['Requerimentos-Análise']
    tabRequerimentos = ws.tables['tabRequerimentos']
    
    for i, row in enumerate(rows, start=1):
        cols = row.query_selector_all("td") 
        dados = [col.inner_text().strip() for col in cols]
        dados[0] = 'AUTOMATICO'
        
        linhas_existentes = [tuple(row) for row in ws.iter_rows(values_only=True)]
        
        if (dados[9] == 'Em Análise') or (dados[9] == 'Em Análise - RE'):
            novalinha = dados[:10]
            if tuple(novalinha) not in linhas_existentes:
                ws.append(novalinha)
                print(f"   ✅ Linha adicionada: {novalinha[1]}")
                novosAnalises.append(novalinha[1])
            else:
                print(f"   ⏭️  Linha já existe: {novalinha[1]}")
    
    # Atualiza referência da tabela
    ref_atual = tabRequerimentos.ref
    ultima_linha = ws.max_row
    col_inicio, linha_inicio = ref_atual.split(":")[0][0], ref_atual.split(":")[0][1:]
    col_fim = ref_atual.split(":")[1][0]
    nova_ref = f"{col_inicio}{linha_inicio}:{col_fim}{ultima_linha}"
    tabRequerimentos.ref = nova_ref

    wb.save(LOG_REQUERIMENTOS)
    wb.close()
    
    print(f"\n📊 Total de novos requerimentos: {len(novosAnalises)}")
    return novosAnalises


def wait_primefaces_ajax(page, timeout=5000):
    """Espera todas as requisições AJAX do PrimeFaces terminarem"""
    try:
        page.wait_for_function(
            """() => {
                if (typeof PrimeFaces === 'undefined') return true;
                if (!PrimeFaces.ajax) return true;
                if (!PrimeFaces.ajax.Queue) return true;
                return PrimeFaces.ajax.Queue.isEmpty();
            }""",
            timeout=timeout
        )
    except:
        pass
    time.sleep(0.3)


def primefaces_click(page, element, description="elemento"):
    """
    Clica em botões submit do PrimeFaces (type="submit").
    Esses botões precisam submeter o formulário via AJAX.
    """
    #print(f"   🎯 Baixando: {description}")
    
    # Scroll até o elemento
    try:
        element.scroll_into_view_if_needed()
        time.sleep(0.3)
    except:
        pass
    
    # MÉTODO 1: Executa o onclick diretamente se existir
    try:
        onclick_executed = page.evaluate("""(button) => {
            const onclick = button.getAttribute('onclick');
            if (onclick) {
                try {
                    // Executa o código onclick
                    eval(onclick);
                    return true;
                } catch (e) {
                    console.error('Erro ao executar onclick:', e);
                    return false;
                }
            }
            return false;
        }""", element)
        
        if onclick_executed:
            print(f"   ✅ Onclick executado diretamente")
            time.sleep(1)
            return True
    except Exception as e:
        print(f"   ⚠ Onclick falhou: {str(e)[:50]}")
    
    # MÉTODO 2: Submit via PrimeFaces.ajax.Request
    try:
        success = page.evaluate("""(button) => {
            try {
                const form = button.closest('form');
                if (!form) return false;
                
                // Cria input hidden com dados do botão
                const hiddenInput = document.createElement('input');
                hiddenInput.type = 'hidden';
                hiddenInput.name = button.name || button.id;
                hiddenInput.value = button.value || button.id;
                form.appendChild(hiddenInput);
                
                // Usa PrimeFaces.ajax.Request se disponível
                if (typeof PrimeFaces !== 'undefined' && PrimeFaces.ajax && PrimeFaces.ajax.Request) {
                    const options = {
                        source: button.id,
                        process: button.id,
                        update: '@form',
                        formId: form.id,
                        params: []
                    };
                    
                    const paramObj = {};
                    paramObj[button.name || button.id] = button.value || button.id;
                    options.params.push(paramObj);
                    
                    PrimeFaces.ajax.Request.handle(options);
                    return true;
                }
                
                // Fallback: dispara submit
                const submitEvent = new Event('submit', {
                    bubbles: true,
                    cancelable: true
                });
                
                form._submitButton = button;
                form.dispatchEvent(submitEvent);
                
                if (!submitEvent.defaultPrevented) {
                    if (typeof jsf !== 'undefined' && jsf.ajax && jsf.ajax.request) {
                        jsf.ajax.request(button, null, {
                            'javax.faces.behavior.event': 'action'
                        });
                        return true;
                    }
                    form.submit();
                }
                
                return true;
            } catch (e) {
                console.error('Erro ao submeter:', e);
                return false;
            }
        }""", element)
        
        if success:
            print(f"   ✅ Aguardando resposta do Mosaico...")
            time.sleep(1)
            return True
        else:
            print(f"   ⚠ Submit falhou")
    except Exception as e:
        print(f"   ❌ Erro: {str(e)[:80]}")
    
    # MÉTODO 3: Force click como último recurso
    try:
        print(f"   🔄 Tentando force click...")
        element.click(force=True, timeout=2000)
        print(f"   ✅ Force click funcionou")
        time.sleep(1)
        return True
    except Exception as e:
        print(f"   ❌ Force click falhou: {str(e)[:50]}")
    
    return False


def baixar_pdfs(page, requerimento):
    """Baixa todos os PDFs da página de anexos"""
    num, ano = requerimento.split("/")
    pasta_destino = os.path.join(FILES_FOLDER, f"Requerimentos\\_{ano}.{num}")
    
    # Cria a pasta do requerimento se não existir
    pasta_destino = criar_pasta_se_nao_existir(requerimento)
    
    #print(f"   📥 Buscando PDFs para baixar...")
    
    # Lista de botões que revelam PDFs
    botoes_pdf = [
        "Outros",
        "ART", 
        "Selo ANATEL",
        "Relatório de Avaliação da Conformidade - RACT",
        "Manual do Produto",
        "Certificado de Conformidade Técnica - CCT",
        "Contrato Social",
        "Fotos internas",
        "Relatório de Ensaio",
        "Fotos do produto"
    ]
    
    total_pdfs_baixados = 0
    
    # Percorre cada botão para revelar PDFs
    for nome_botao in botoes_pdf:
        #print(f"   🔍 Procurando botão: {nome_botao}")
        
        try:
            # Busca o botão pelo nome/texto
            botao = page.get_by_role("button", name=nome_botao)
            
            # Verifica se o botão existe e está visível
            if botao.count() > 0:
                print(f"   🎯 Buscando: {nome_botao}")
                
                # Clica no botão
                botao.first.click()
                
                # Aguarda o carregamento dos PDFs
                time.sleep(1)
                wait_primefaces_ajax(page)
                
                # Busca todos os links de PDF que foram revelados
                pdf_links = page.query_selector_all("a[href*='.pdf'], a[href*='download']")

                # Extrai informações da tabela
                tabela = page.query_selector("table.analiseTable")
                linhas_dados = []
                
                if tabela:
                    # Pega todas as linhas da tabela (exceto cabeçalho)
                    linhas = tabela.query_selector_all("tr")
                    
                    # Identifica o cabeçalho (primeira linha)
                    if len(linhas) > 0:
                        cabecalho = linhas[0].query_selector_all("th, td")
                        headers = [col.inner_text().strip() for col in cabecalho]
                        #print(f"   📋 Cabeçalho da tabela: {headers}")
                        
                        # Processa as linhas de dados (exceto cabeçalho)
                        for linha in linhas[1:]:
                            colunas = linha.query_selector_all("th, td")
                            dados = [col.inner_text().strip() for col in colunas]
                            
                            if len(dados) >= len(headers):
                                linha_info = {}
                                for i, header in enumerate(headers):
                                    if i < len(dados):
                                        linha_info[header] = dados[i]
                                linhas_dados.append(linha_info)
                    
                    #print(f"   📊 {len(linhas_dados)} linha(s) de dados extraída(s)")
                
                if pdf_links:
                    print(f"   📄 {len(pdf_links)} PDF(s) encontrado(s) para {nome_botao}")
                    
                    # Verifica se há correspondência entre PDFs e linhas da tabela
                    if len(pdf_links) != len(linhas_dados):
                        print(f"   ⚠ AVISO: {len(pdf_links)} PDFs mas {len(linhas_dados)} linhas na tabela!")
                    
                    # Baixa cada PDF encontrado
                    for idx, link in enumerate(pdf_links):
                        try:
                            # Verifica se o link está visível/disponível
                            if not link.is_visible():
                                continue
                            
                            # Pega informações da linha correspondente da tabela
                            linha_info = None
                            if idx < len(linhas_dados):
                                linha_info = linhas_dados[idx]
                            
                            # Pega informações da linha correspondente da tabela para gerar nome do arquivo
                            linha_info = None
                            if idx < len(linhas_dados):
                                linha_info = linhas_dados[idx]
                            
                            # Pega o nome do arquivo original do link
                            href = link.get_attribute('href')
                            texto = link.inner_text().strip()
                            
                            # Gera nome do arquivo baseado nas informações da tabela (ANTES do download)
                            nome_arquivo_temp = f"anexo_{idx + 1}.pdf"  # Nome temporário padrão
                            
                            if linha_info:
                                try:
                                    # Extrai informações da tabela
                                    doc_id = linha_info.get("ID", f"#{idx + 1}")
                                    tipo_doc = linha_info.get("Tipo de Documento", "Documento")
                                    data_hora = linha_info.get("Data - Hora", "")
                                    
                                    # Processa a data para formato yyyy.mm.dd
                                    data_formatada = ""
                                    if data_hora:
                                        try:
                                            # Remove espaços extras e separa data da hora
                                            data_parte = data_hora.split()[0] if data_hora else ""
                                            
                                            # Padrões de data comuns
                                            padroes = [
                                                r"(\d{2})/(\d{2})/(\d{4})",  # dd/mm/yyyy
                                                r"(\d{4})-(\d{2})-(\d{2})",  # yyyy-mm-dd
                                                r"(\d{2})-(\d{2})-(\d{4})",  # dd-mm-yyyy
                                            ]
                                            
                                            for padrao in padroes:
                                                match = re.search(padrao, data_parte)
                                                if match:
                                                    if padrao == padroes[0] or padrao == padroes[2]:  # dd/mm/yyyy ou dd-mm-yyyy
                                                        dia, mes, ano = match.groups()
                                                    else:  # yyyy-mm-dd
                                                        ano, mes, dia = match.groups()
                                                    
                                                    data_formatada = f"{ano}.{mes.zfill(2)}.{dia.zfill(2)}"
                                                    break
                                            
                                            if not data_formatada:
                                                data_formatada = "0000.00.00"
                                                print(f"   ⚠ Não foi possível processar a data: {data_hora}")
                                        except Exception as e:
                                            data_formatada = "0000.00.00"
                                            print(f"   ⚠ Erro ao processar data: {str(e)[:50]}")
                                    else:
                                        data_formatada = "0000.00.00"
                                    
                                    # Limpa caracteres inválidos para nome de arquivo
                                    tipo_doc_limpo = re.sub(r'[<>:"/\\|?*]', '_', tipo_doc)
                                    doc_id_limpo = re.sub(r'[<>:"/\\|?*]', '_', str(doc_id))
                                    
                                    # Monta o novo nome: [tipo][data - ID id] nome_original [req num de ano].ext
                                    nome_arquivo_temp = f"[{tipo_doc_limpo}][{data_formatada} - ID {doc_id_limpo}] temp [req {num} de {ano}].pdf"
                                    
                                except Exception as e:
                                    print(f"   ⚠ Erro ao processar informações da tabela: {str(e)[:50]}")
                                    # Fallback para nome simples
                                    nome_arquivo_temp = f"[{nome_botao}] anexo_{idx + 1}.pdf"
                            else:
                                # Fallback quando não há informação da tabela
                                nome_arquivo_temp = f"[{nome_botao}] anexo_{idx + 1}.pdf"
                            
                            # Verifica se arquivo já existe ANTES de iniciar o download
                            caminho_completo_temp = os.path.join(pasta_destino, nome_arquivo_temp)
                            
                            # Procura por arquivos existentes que começem com o mesmo padrão
                            nome_base_busca = nome_arquivo_temp.split('] temp [')[0] + ']'  # Remove "temp" e tudo depois
                            arquivos_existentes = [f for f in os.listdir(pasta_destino) if f.startswith(nome_base_busca)]
                            
                            if arquivos_existentes:
                                print(f"   ⏭️ Arquivo já existe, pulando: {arquivos_existentes[0]}")
                                continue
                            
                            # Se não existe, faz o download
                            with page.expect_download() as download_info:
                                link.click()
                            
                            download = download_info.value
                            
                            # Pega o nome real do arquivo baixado
                            nome_arquivo_real = download.suggested_filename
                            if not nome_arquivo_real or nome_arquivo_real == "":
                                nome_arquivo_real = f"anexo_{idx + 1}.pdf"
                            
                            # Extrai nome base e extensão do arquivo real
                            nome_base_real, extensao_real = os.path.splitext(nome_arquivo_real)
                            
                            # Monta o nome final usando o nome real do arquivo
                            if linha_info:
                                try:
                                    doc_id = linha_info.get("ID", f"#{idx + 1}")
                                    tipo_doc = linha_info.get("Tipo de Documento", "Documento")
                                    data_hora = linha_info.get("Data - Hora", "")
                                    
                                    # Usa a mesma lógica de formatação de data
                                    data_formatada = "0000.00.00"
                                    if data_hora:
                                        data_parte = data_hora.split()[0] if data_hora else ""
                                        padroes = [
                                            r"(\d{2})/(\d{2})/(\d{4})",
                                            r"(\d{4})-(\d{2})-(\d{2})",
                                            r"(\d{2})-(\d{2})-(\d{4})",
                                        ]
                                        for padrao in padroes:
                                            match = re.search(padrao, data_parte)
                                            if match:
                                                if padrao == padroes[0] or padrao == padroes[2]:
                                                    dia, mes, ano = match.groups()
                                                else:
                                                    ano, mes, dia = match.groups()
                                                data_formatada = f"{ano}.{mes.zfill(2)}.{dia.zfill(2)}"
                                                break
                                    
                                    tipo_doc_limpo = re.sub(r'[<>:"/\\|?*]', '_', tipo_doc)
                                    doc_id_limpo = re.sub(r'[<>:"/\\|?*]', '_', str(doc_id))
                                    
                                    nome_arquivo_final = f"[{tipo_doc_limpo}][{data_formatada} - ID {doc_id_limpo}] {nome_base_real} [req {num} de {ano}]{extensao_real}"
                                except:
                                    nome_arquivo_final = f"[{nome_botao}] {nome_base_real}{extensao_real}"
                            else:
                                nome_arquivo_final = f"[{nome_botao}] {nome_base_real}{extensao_real}"
                            
                            # Salva o arquivo com o nome final
                            caminho_completo = os.path.join(pasta_destino, nome_arquivo_final)
                            download.save_as(caminho_completo)
                            
                            print(f"   ✅ Baixado: {nome_arquivo_final}")
                            total_pdfs_baixados += 1
                            
                        except Exception as e:
                            print(f"   ❌ Erro ao baixar PDF {idx + 1} de {nome_botao}: {str(e)[:50]}")
                else:
                    print(f"   ℹ Nenhum PDF encontrado para: {nome_botao}")
                    
            else:
                print(f"   ⚠ Botão não encontrado: {nome_botao}")
                
        except Exception as e:
            print(f"   ❌ Erro ao processar botão {nome_botao}: {str(e)[:50]}")
            continue
    
    if total_pdfs_baixados == 0:
        print(f"   ⚠ Nenhum PDF foi baixado")
    else:
        print(f"   💾 Total de {total_pdfs_baixados} PDF(s) salvos em: {pasta_destino}")

def abrir_caixa_de_entrada(page_obj):
    # Navega para a lista
    lista_url = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"
    page_obj.goto(lista_url)
    page_obj.wait_for_load_state("load")
    
    # Clica em "Todos"
    page_obj.click("#menuForm\\:todos",timeout=3600000) 
    page_obj.wait_for_load_state("load")
    
    # Seleciona 100 itens por página e aguarda atualização
    page_obj.select_option("select.ui-paginator-rpp-options", value="100")
    page_obj.wait_for_load_state("networkidle")  # Aguarda requisições AJAX terminarem
    wait_primefaces_ajax(page_obj)
    time.sleep(1)  # Pausa adicional para garantir que a tabela seja recarregada
    
    return page_obj

with sync_playwright() as p:
    browser = p.chromium.launch_persistent_context(
        PROFILE_DIR,
        headless=False,
        executable_path=CHROME_PATH,
        args=[
            "--start-maximized",
            "--disable-blink-features=AutomationControlled"
        ],
        accept_downloads=True  # IMPORTANTE: permite downloads
    )
    
    page = browser.new_page()
    
    # Inicia timer para controle de timeout do Mosaico
    import time
    inicio_execucao = time.time()
    limite_tempo = 28 * 60  # 28 minutos em segundos
    
    # Navega para a lista
    page = abrir_caixa_de_entrada(page)
    '''lista_url = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"
    page.goto(lista_url)
    page.wait_for_load_state("load")
    
    # Clica em "Todos"
    page.click("#menuForm\\:todos",timeout=3600000) 
    page.wait_for_load_state("load")
    wait_primefaces_ajax(page)'''
    
    print("\n" + "="*60)
    print("🤖 AUTOMAÇÃO ORCN - DOWNLOAD DE ANEXOS")
    print("="*60)
    
    #input("➡ Faça login (se necessário) e pressione ENTER para continuar...")
    
    # Seleciona todas as linhas
    rows = page.query_selector_all("css=#form\\:tarefasTable_data tr")
    print(f"\n🔎 {len(rows)} linhas encontradas na tabela")
    
    # Atualiza Excel (modo debug) ou pega todos os requerimentos (modo produção)
    if debug_mode:
        print("\n📊 Atualizando planilha Excel...")
    else:
        print("\n📋 Coletando requerimentos da lista...")
    
    novos_requerimentos = atualizar_excel(rows)
    
    if not novos_requerimentos:
        if debug_mode:
            print("\n✅ Nenhum requerimento novo para processar!")
        else:
            print("\n✅ Nenhum requerimento encontrado na lista!")
        input("Pressione ENTER para encerrar...")
        browser.close()
        exit()
    
    if debug_mode:
        print(f"\n🚀 Processando {len(novos_requerimentos)} requerimento(s) novo(s)...\n")
    else:
        print(f"\n🚀 Processando {len(novos_requerimentos)} requerimento(s) da lista...\n")
    
    # Cria um dicionário com os dados de cada linha ANTES de iterar
    print("\n📋 Mapeando linhas da tabela...")
    linhas_dados = []
    
    for i, row in enumerate(reversed(rows), start=1):
        try:
            cols = row.query_selector_all("td")
            if len(cols) < 2:
                continue
            
            requerimento = cols[1].inner_text().strip()
            
            # Armazena os dados da linha
            linhas_dados.append({
                'indice': i,
                'requerimento': requerimento,
                'row': row
            })
        except Exception as e:
            print(f"   ⚠ Erro ao ler linha {i}: {str(e)[:50]}")
    
    print(f"   ✅ {len(linhas_dados)} linhas mapeadas")
    
    # Processa cada linha dos dados salvos
    for linha_info in linhas_dados:
        # Verifica se o tempo limite foi atingido (27 minutos)
        tempo_decorrido = time.time() - inicio_execucao
        if tempo_decorrido > limite_tempo:
            minutos_decorridos = int(tempo_decorrido // 60)
            print(f"\n{'='*60}")
            print(f"⏰ TIMEOUT PREVENTIVO ATIVADO!")
            print(f"{'='*60}")
            print(f"⚠ Tempo decorrido: {minutos_decorridos} minutos")
            print(f"⚠ Encerrando aplicação para evitar timeout do Mosaico (30 min)")
            print(f"⚠ Execute novamente o script para continuar processando")
            print(f"{'='*60}")
            input("\nPressione ENTER para encerrar...")
            browser.close()
            exit()
        
        i = linha_info['indice']
        requerimento = linha_info['requerimento']
        
        # Só processa se for um requerimento novo
        full_path = req_para_fullpath(requerimento)
        if (requerimento not in novos_requerimentos) or (os.path.exists(full_path)):
            continue
        
        print(f"\n{'='*60}")
        print(f"▶ Requerimento {i}: {requerimento}")
        print(f"{'='*60}")
        
        # IMPORTANTE: Recarrega a linha atual usando busca manual
        # Busca pela linha que contém este requerimento específico
        row_atual = None
        try:
            # Recarrega todas as linhas da tabela
            linhas_atualizadas = page.query_selector_all("css=#form\\:tarefasTable_data tr")
            #print(f"   🔄 Recarregadas {len(linhas_atualizadas)} linhas da tabela")
            
            # Procura manualmente pela linha com o requerimento
            for idx, linha in enumerate(linhas_atualizadas):
                try:
                    cols = linha.query_selector_all("td")
                    if len(cols) >= 2:
                        requerimento_da_linha = cols[1].inner_text().strip()
                        if requerimento_da_linha == requerimento:
                            row_atual = linha
                            #print(f"   ✅ Requerimento {requerimento} encontrado na posição {idx + 1}")
                            break
                except:
                    continue
            
            if not row_atual:
                print(f"   ⚠ Requerimento {requerimento} não encontrado na lista atualizada, pulando...")
                continue
                
        except Exception as e:
            print(f"   ⚠ Erro ao recarregar linhas: {str(e)[:50]}, pulando...")
            continue
        
        # Busca botão "Visualizar em Tela cheia" na linha atual 
        btn = row_atual.query_selector("button[type='submit'][title='Visualizar em Tela cheia']")
        if not btn:
            btn = row_atual.query_selector("button[title*='Tela cheia']")
        
        if not btn:
            print(f"   ⚠ Botão não encontrado, pulando...")
            continue
        
        # Clica no botão
        if not primefaces_click(page, btn, "Visualizar em Tela cheia"):
            print(f"   ⚠ Não foi possível clicar, pulando...")
            continue
        
        # Aguarda carregar
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except:
            pass
        wait_primefaces_ajax(page)
        
        print(f"   🌐 URL: {page.url}")
        
        # Busca botão "Anexos"
        time.sleep(2)
        #print(f"   Buscando botão Anexos na toolbar...")
        
        iframe_element = page.wait_for_selector("#__frameDetalhe", timeout=10000)

        if iframe_element:
            detalhes_requerimento = iframe_element.get_attribute("src")
            #print(detalhes_requerimento)
            if detalhes_requerimento:
                page.goto(detalhes_requerimento)
                anexos_btn = page.get_by_role("button", name="Anexos")
                try:
                    #anexos_btn = page.query_selector("div.ui-toolbar-group-left button:has-text('Anexos')")
                    if anexos_btn:                        
                        anexos_btn.click(no_wait_after=True)
                        # Aguarda um seletor específico da nova “tela” ou da área que muda
                        page.wait_for_selector(".ui-blockui", state="detached", timeout=15000)
                        print("   🔄 Buscando anexos...")
                    else:
                        print("Botão 'Anexos' não encontrado.")
                    wait_primefaces_ajax(page)
                        
                    print(f"   ✅ Página de Anexos carregada")
                        
                        # BAIXA OS PDFs
                    baixar_pdfs(page, requerimento)                        
                except Exception as e:
                    print(f"   ⚠ expect_navigation falhou: {str(e)}")
        else:
            print("⚠ iframe_element não encontrado, pulando...")
            continue

        #anexos_btn = None

        #primefaces_click
        #teogenes

        # Método 0: Botão com span contendo "Anexos"
        '''try:
            #anexos_btn = page.get_by_role("button", name="Anexos")
            #anexos_btn = page.query_selector("#formAnalise\\:j_idt110\\:2\\:j_idt115")
            #anexos_btn = page.query_selector("xpath=/html/body/span/form/span/div/div/div[1]/span/div/div/span[3]/span/button")
            #anexos_btn = page.query_selector("button[type='submit'][title='Anexos']")
            
            if anexos_btn:
                try:
                    btn_id = page.evaluate("(btn) => btn.id", anexos_btn)
                    print(f"   ID do botão: {btn_id}")
                except:
                    pass
                if primefaces_click(page, anexos_btn, "Anexos"):
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except:
                        pass
                    wait_primefaces_ajax(page)
                    
                    print(f"   ✅ Página de Anexos carregada")
                    
                    # BAIXA OS PDFs
                    baixar_pdfs(page, requerimento)
                else:
                    print(f"   ⚠ Não foi possível clicar em Anexos")
            else:
                print(f"   ⚠ Botão 'Anexos' não encontrado")                    
        except:
            pass'''
        
        # Volta para a lista
        print(f"   ↩ Voltando para a lista...")
        page = abrir_caixa_de_entrada(page)
        '''page.goto(lista_url, wait_until="networkidle")
        wait_primefaces_ajax(page)
        time.sleep(1)  # Pequena pausa para garantir que o DOM esteja estável'''
    
    print("\n" + "="*60)
    print("✅ PROCESSAMENTO CONCLUÍDO!")
    print("="*60)
    input("\nPressione ENTER para encerrar...")
    browser.close()