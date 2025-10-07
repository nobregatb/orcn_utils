from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os
import time

CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
FILES_FOLDER = r"C:\Users\tbnobrega\OneDrive - ANATEL\Anatel\_ORCN"
PROFILE_DIR = os.path.abspath("./meu_perfil_chrome")
LOG_REQUERIMENTOS = os.path.join(FILES_FOLDER, 'ORCN - Documentação pertinente.xlsx')


def criar_pasta_se_nao_existir(req):
    """Cria pasta do requerimento se não existir"""
    num, ano = req.split("/")
    requerimento = rf"Requerimentos\{ano}.{num}"
    full_path = os.path.join(FILES_FOLDER, requerimento)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
        print(f"   📁 Pasta criada: {full_path}")
    return full_path


def atualizar_excel(rows) -> list:
    """Atualiza planilha com novos requerimentos em análise"""
    novosAnalises = []
    wb = load_workbook(LOG_REQUERIMENTOS)
    ws = wb['Requerimentos-Análise']
    tabRequerimentos = ws.tables['tabRequerimentos']
    
    for i, row in enumerate(rows, start=1):
        cols = row.query_selector_all("td") 
        dados = [col.inner_text().strip() for col in cols]
        dados[0] = 'AUTOMATICO'
        
        linhas_existentes = [tuple(row) for row in ws.iter_rows(values_only=True)]
        
        if dados[9] == 'Em Análise':
            novalinha = dados[:10]
            if tuple(novalinha) not in linhas_existentes:
                ws.append(novalinha)
                print(f"   ✅ Linha adicionada: {novalinha[1]}")
                novosAnalises.append(novalinha[1])
                criar_pasta_se_nao_existir(novalinha[1])
            else:
                print(f"   ⏭️  Linha já existe: {novalinha[1]}")
    
    # Atualiza referência da tabela
    ref_atual = tabRequerimentos.ref
    ultima_linha = ws.max_row
    col_inicio, linha_inicio = ref_atual.split(":")[0][0], ref_atual.split(":")[0][1:]
    col_fim = ref_atual.split(":")[1][0]
    nova_ref = f"{col_inicio}{linha_inicio}:{col_fim}{ultima_linha}"
    tabRequerimentos.ref = nova_ref

    wb.save(LOG_REQUERIMENTOS)
    wb.close()
    
    print(f"\n📊 Total de novos requerimentos: {len(novosAnalises)}")
    return novosAnalises


def wait_primefaces_ajax(page, timeout=5000):
    """Espera todas as requisições AJAX do PrimeFaces terminarem"""
    try:
        page.wait_for_function(
            """() => {
                if (typeof PrimeFaces === 'undefined') return true;
                if (!PrimeFaces.ajax) return true;
                if (!PrimeFaces.ajax.Queue) return true;
                return PrimeFaces.ajax.Queue.isEmpty();
            }""",
            timeout=timeout
        )
    except:
        pass
    time.sleep(0.3)


def primefaces_click(page, element, description="elemento"):
    """
    Clica em botões submit do PrimeFaces (type="submit").
    Esses botões precisam submeter o formulário via AJAX.
    """
    print(f"   🎯 Clicando em: {description}")
    
    # Scroll até o elemento
    try:
        element.scroll_into_view_if_needed()
        time.sleep(0.3)
    except:
        pass
    
    # MÉTODO ESPECÍFICO PARA BOTÕES SUBMIT DO PRIMEFACES
    try:
        success = page.evaluate("""(button) => {
            try {
                const form = button.closest('form');
                if (!form) return false;
                
                // Cria input hidden com dados do botão
                const hiddenInput = document.createElement('input');
                hiddenInput.type = 'hidden';
                hiddenInput.name = button.name || button.id;
                hiddenInput.value = button.value || button.id;
                form.appendChild(hiddenInput);
                
                // Usa PrimeFaces.ajax.Request se disponível
                if (typeof PrimeFaces !== 'undefined' && PrimeFaces.ajax && PrimeFaces.ajax.Request) {
                    const options = {
                        source: button.id,
                        process: button.id,
                        update: '@form',
                        formId: form.id,
                        params: []
                    };
                    
                    const paramObj = {};
                    paramObj[button.name || button.id] = button.value || button.id;
                    options.params.push(paramObj);
                    
                    PrimeFaces.ajax.Request.handle(options);
                    return true;
                }
                
                // Fallback: dispara submit
                const submitEvent = new Event('submit', {
                    bubbles: true,
                    cancelable: true
                });
                
                form._submitButton = button;
                form.dispatchEvent(submitEvent);
                
                if (!submitEvent.defaultPrevented) {
                    if (typeof jsf !== 'undefined' && jsf.ajax && jsf.ajax.request) {
                        jsf.ajax.request(button, null, {
                            'javax.faces.behavior.event': 'action'
                        });
                        return true;
                    }
                    form.submit();
                }
                
                return true;
            } catch (e) {
                console.error('Erro ao submeter:', e);
                return false;
            }
        }""", element)
        
        if success:
            print(f"   ✅ Submit executado")
            time.sleep(1)
            return True
        else:
            print(f"   ⚠ Submit falhou")
    except Exception as e:
        print(f"   ❌ Erro: {str(e)[:80]}")
    
    # Fallback: force click
    try:
        print(f"   🔄 Tentando force click...")
        element.click(force=True, timeout=2000)
        print(f"   ✅ Force click funcionou")
        time.sleep(1)
        return True
    except Exception as e:
        print(f"   ❌ Force click falhou: {str(e)[:50]}")
    
    return False


def baixar_pdfs(page, requerimento):
    """Baixa todos os PDFs da página de anexos"""
    num, ano = requerimento.split("/")
    pasta_destino = os.path.join(FILES_FOLDER, f"Requerimentos\\{ano}.{num}")
    
    print(f"   📥 Buscando PDFs para baixar...")
    
    # Busca todos os links de PDF
    pdf_links = page.query_selector_all("a[href*='.pdf'], a[href*='download']")
    
    if not pdf_links:
        print(f"   ⚠ Nenhum PDF encontrado")
        return
    
    print(f"   📄 {len(pdf_links)} PDF(s) encontrado(s)")
    
    for idx, link in enumerate(pdf_links, start=1):
        try:
            # Pega o nome do arquivo
            href = link.get_attribute('href')
            texto = link.inner_text().strip()
            
            # Configura download
            with page.expect_download() as download_info:
                link.click()
            
            download = download_info.value
            
            # Define nome do arquivo
            nome_arquivo = download.suggested_filename
            if not nome_arquivo or nome_arquivo == "":
                nome_arquivo = f"anexo_{idx}.pdf"
            
            # Salva o arquivo
            caminho_completo = os.path.join(pasta_destino, nome_arquivo)
            download.save_as(caminho_completo)
            
            print(f"   ✅ Baixado: {nome_arquivo}")
            
        except Exception as e:
            print(f"   ❌ Erro ao baixar PDF {idx}: {str(e)[:50]}")
    
    print(f"   💾 PDFs salvos em: {pasta_destino}")


with sync_playwright() as p:
    browser = p.chromium.launch_persistent_context(
        PROFILE_DIR,
        headless=False,
        executable_path=CHROME_PATH,
        args=[
            "--start-maximized",
            "--disable-blink-features=AutomationControlled"
        ],
        accept_downloads=True  # IMPORTANTE: permite downloads
    )
    
    page = browser.new_page()
    
    # Navega para a lista
    lista_url = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"
    page.goto(lista_url)
    page.wait_for_load_state("load")
    
    # Clica em "Todos"
    page.click("#menuForm\\:todos") 
    page.wait_for_load_state("load")
    wait_primefaces_ajax(page)
    
    print("\n" + "="*60)
    print("🤖 AUTOMAÇÃO ORCN - DOWNLOAD DE ANEXOS")
    print("="*60)
    
    input("➡ Faça login (se necessário) e pressione ENTER para continuar...")
    
    # Seleciona todas as linhas
    rows = page.query_selector_all("css=#form\\:tarefasTable_data tr")
    print(f"\n🔎 {len(rows)} linhas encontradas na tabela")
    
    # Atualiza Excel e pega novos requerimentos
    print("\n📊 Atualizando planilha Excel...")
    novos_requerimentos = atualizar_excel(rows)
    
    if not novos_requerimentos:
        print("\n✅ Nenhum requerimento novo para processar!")
        input("Pressione ENTER para encerrar...")
        browser.close()
        exit()
    
    print(f"\n🚀 Processando {len(novos_requerimentos)} requerimento(s) novo(s)...\n")
    
    # Processa cada linha
    for i, row in enumerate(rows, start=1):
        # Pega o número do requerimento desta linha
        cols = row.query_selector_all("td")
        if len(cols) < 2:
            continue
        
        requerimento = cols[1].inner_text().strip()
        
        # Só processa se for um requerimento novo
        if requerimento not in novos_requerimentos:
            continue
        
        print(f"\n{'='*60}")
        print(f"▶ LINHA {i}: {requerimento}")
        print(f"{'='*60}")
        
        # Busca botão "Visualizar em Tela cheia"
        btn = row.query_selector("button[type='submit'][title='Visualizar em Tela cheia']")
        if not btn:
            btn = row.query_selector("button[title*='Tela cheia']")
        
        if not btn:
            print(f"   ⚠ Botão não encontrado, pulando...")
            continue
        
        # Clica no botão
        if not primefaces_click(page, btn, "Visualizar em Tela cheia"):
            print(f"   ⚠ Não foi possível clicar, pulando...")
            continue
        
        # Aguarda carregar
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except:
            pass
        wait_primefaces_ajax(page)
        
        print(f"   🌐 URL: {page.url}")
        
        # Busca botão "Anexos"
        time.sleep(1)
        anexos_btn = (
            page.query_selector("//button[.//span[contains(text(),'Anexos')]]") or
            page.query_selector("button:has-text('Anexos')") or
            page.query_selector("button[type='submit'][title*='Anexo' i]")
        )
        
        if anexos_btn:
            if primefaces_click(page, anexos_btn, "Anexos"):
                try:
                    page.wait_for_load_state("networkidle", timeout=10000)
                except:
                    pass
                wait_primefaces_ajax(page)
                
                print(f"   ✅ Página de Anexos carregada")
                
                # BAIXA OS PDFs
                baixar_pdfs(page, requerimento)
            else:
                print(f"   ⚠ Não foi possível clicar em Anexos")
        else:
            print(f"   ⚠ Botão 'Anexos' não encontrado")
        
        # Volta para a lista
        print(f"   ↩ Voltando para a lista...")
        page.goto(lista_url, wait_until="networkidle")
        wait_primefaces_ajax(page)
        
        # Recarrega as linhas
        rows = page.query_selector_all("css=#form\\:tarefasTable_data tr")
    
    print("\n" + "="*60)
    print("✅ PROCESSAMENTO CONCLUÍDO!")
    print("="*60)
    input("\nPressione ENTER para encerrar...")
    browser.close()