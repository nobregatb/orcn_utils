# -*- coding: utf-8 -*-
"""
Funções utilitárias centralizadas do projeto ORCN Utils.
Contém todas as funções auxiliares e utilitárias usadas em múltiplos módulos.
"""

import os
import sys
import re
import json
import subprocess
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import pandas as pd
from openpyxl import load_workbook
from core.const import EXCEL_PATH, REQUERIMENTOS_PATH, TAB_REQUERIMENTOS, EXCEL_SHEET_NAME, DOWNLOAD_LOG_FILENAME
from core.log_print import log_info, log_erro
# Imports opcionais para funcionalidades específicas
try:
    from pdf2image import convert_from_path
    import pytesseract
    OCR_DISPONIVEL = True
except ImportError:
    OCR_DISPONIVEL = False

from core.const import (
    TBN_FILES_FOLDER, CHROME_PROFILE_DIR, REQUERIMENTOS_DIR_INBOX,
    GIT_COMMANDS, GIT_TIMEOUT, VERSAO_PADRAO, MENSAGENS_STATUS, TIPOS_DOCUMENTOS,
    TESSERACT_PATH, JSON_FILES
)
from core.log_print import log_info, log_erro, log_erro_critico

# ================================
# FUNÇÕES DE VALIDAÇÃO CRÍTICA  
# ================================

def validar_dados_criticos(requerimento_json=None, dados_ocd=None, dados_lab=None, 
                         dados_fabricante=None, dados_solicitante=None, 
                         nome_requerimento="", contexto=""):
    """
    Valida se os dados críticos foram lidos corretamente e para a aplicação se necessário.
    
    Args:
        requerimento_json: Dados do requerimento extraídos
        dados_ocd: Dados do OCD extraídos  
        dados_lab: Dados do laboratório extraídos
        dados_fabricante: Dados do fabricante extraídos
        dados_solicitante: Dados do solicitante extraídos
        nome_requerimento: Nome do requerimento para identificação
        contexto: Contexto da operação (download, análise, etc.)
    
    Raises:
        SystemExit: Se dados críticos não forem válidos
    """
    #log_info(f"🔍 Validando dados críticos do requerimento {nome_requerimento} ({contexto})")
    
    erros_criticos = []
    
    # Validar requerimento_json
    if requerimento_json is not None:
        if not isinstance(requerimento_json, dict) or not requerimento_json:
            erros_criticos.append("requerimento_json não é um dicionário válido ou está vazio")
        elif not requerimento_json.get('num_req'):
            erros_criticos.append("requerimento_json não contém número do requerimento (num_req)")
    
    # Validar dados_ocd
    if dados_ocd is not None:
        if not isinstance(dados_ocd, dict):
            erros_criticos.append("dados_ocd não é um dicionário válido")
        elif not dados_ocd:
            erros_criticos.append("dados_ocd está vazio - dados do OCD são obrigatórios")
        elif not dados_ocd.get('CNPJ') and not dados_ocd.get('Nome'):
            erros_criticos.append("dados_ocd não contém CNPJ nem Nome do OCD")
    
    # Validar dados_lab  
    if dados_lab is not None:
        if not isinstance(dados_lab, dict):
            erros_criticos.append("dados_lab não é um dicionário válido")
        elif not dados_lab:
            erros_criticos.append("dados_lab está vazio - dados do laboratório são obrigatórios")
        elif not dados_lab.get('Nome') and not dados_lab.get('CNPJ'):
            erros_criticos.append("dados_lab não contém Nome nem CNPJ do laboratório")
    
    # Validar dados_fabricante
    if dados_fabricante is not None:
        if not isinstance(dados_fabricante, dict):
            erros_criticos.append("dados_fabricante não é um dicionário válido")
        elif not dados_fabricante:
            erros_criticos.append("dados_fabricante está vazio - dados do fabricante são obrigatórios")
        elif not dados_fabricante.get('Nome') and not dados_fabricante.get('CNPJ'):
            erros_criticos.append("dados_fabricante não contém Nome nem CNPJ do fabricante")
    
    # Validar dados_solicitante
    if dados_solicitante is not None:
        if not isinstance(dados_solicitante, dict):
            erros_criticos.append("dados_solicitante não é um dicionário válido")
        elif not dados_solicitante:
            erros_criticos.append("dados_solicitante está vazio - dados do solicitante são obrigatórios")
        elif not dados_solicitante.get('Nome') and not dados_solicitante.get('CNPJ'):
            erros_criticos.append("dados_solicitante não contém Nome nem CNPJ do solicitante")
    
    # Se há erros críticos, parar a aplicação
    if erros_criticos:
        log_erro_critico(f"❌ ERRO CRÍTICO no requerimento {nome_requerimento} ({contexto}):")
        for erro in erros_criticos:
            log_erro_critico(f"   • {erro}")
        log_erro_critico(f"🛑 APLICAÇÃO INTERROMPIDA - Dados essenciais não foram lidos corretamente!")
        log_erro_critico(f"   Requerimento: {nome_requerimento}")
        log_erro_critico(f"   Contexto: {contexto}")
        sys.exit(1)
    
    #log_info(f"✅ Validação de dados críticos concluída com sucesso para {nome_requerimento}")

# ================================
# FUNÇÕES DE FORMATAÇÃO
# ================================

def limpar_texto(texto, palavras=None, simbolos=None, remover_parenteses=True):
    """
    Remove palavras, símbolos e trechos entre parênteses de uma string.
    Ignora diferenças entre maiúsculas e minúsculas.
    """
    resultado = texto

    # 1️⃣ Remove conteúdo entre parênteses (opcional)
    if remover_parenteses:
        resultado = re.sub(r'\s*\([^)]*\)', '', resultado, flags=re.I)

    # 2️⃣ Remove palavras específicas (case-insensitive)
    if palavras:
        padrao_palavras = r'\b(?:' + '|'.join(map(re.escape, palavras)) + r')\b'
        resultado = re.sub(r'\s*' + padrao_palavras, '', resultado, flags=re.I)

    # 3️⃣ Remove símbolos específicos (case-insensitive, mas não afeta símbolos)
    if simbolos:
        padrao_simbolos = '[' + re.escape(''.join(simbolos)) + ']'
        resultado = re.sub(padrao_simbolos, '', resultado, flags=re.I)

    # 4️⃣ Limpa espaços duplicados e tira espaços extras
    resultado = re.sub(r'\s{2,}', ' ', resultado).strip()

    return resultado



def formatar_cnpj(cnpj_numeros: str) -> str:
    """
    Converte CNPJ de números para formato XX.XXX.XXX/XXXX-XX
    Args:
        cnpj_numeros (str): CNPJ apenas com números (ex: "12345678000123")
    Returns:
        str: CNPJ formatado (ex: "12.345.678/0001-23")
    """
    if not cnpj_numeros or len(cnpj_numeros) != 14:
        return cnpj_numeros
    
    return f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:8]}/{cnpj_numeros[8:12]}-{cnpj_numeros[12:]}"


def desformatar_cnpj(cnpj_formatado: str) -> str:
    """
    Remove formatação do CNPJ, deixando apenas números
    Args:
        cnpj_formatado (str): CNPJ formatado (ex: "12.345.678/0001-23")
    Returns:
        str: CNPJ apenas com números (ex: "12345678000123")
    """
    if not cnpj_formatado:
        return cnpj_formatado
    
    return ''.join(filter(str.isdigit, cnpj_formatado))


def latex_escape_path(caminho: str) -> str:
    """Escapa caracteres especiais do LaTeX dentro de um caminho de arquivo."""
    caminho = caminho.replace("\\", "/")  # usa / para evitar confusão
    # escapa caracteres especiais
    return re.sub(r'([_&#%{}$^~\\])', r'\\\1', caminho)


def escapar_latex(texto: str) -> str:
    """Escapa caracteres especiais do LaTeX."""
    if not isinstance(texto, str):
        return str(texto)
    
    # Dicionário de substituições para caracteres especiais do LaTeX
    substituicoes = {
        '\\': '\\textbackslash{}',
        '{': '\\{',
        '}': '\\}',
        '$': '\\$',
        '&': '\\&',
        '%': '\\%',
        '#': '\\#',
        '^': '\\textasciicircum{}',
        '_': '\\_',
        '~': '\\textasciitilde{}',
    }
    
    # Aplicar substituições
    for char, replacement in substituicoes.items():
        texto = texto.replace(char, replacement)
    
    return texto


# ================================
# FUNÇÕES DE TIPOS DE DOCUMENTO
# ================================

def obter_nome_tipo_documento(tipo_chave: str) -> str:
    """Obtém o nome completo de um tipo de documento pela chave."""
    return TIPOS_DOCUMENTOS.get(tipo_chave, {}).get('nome', 'Tipo desconhecido')

def obter_nome_curto_tipo_documento(tipo_chave: str) -> str:
    """Obtém o nome curto de um tipo de documento pela chave."""
    return TIPOS_DOCUMENTOS.get(tipo_chave, {}).get('nome_curto', 'N/A')

def obter_botao_pdf_tipo_documento(tipo_chave: str) -> str:
    """Obtém o nome do botão PDF para um tipo de documento."""
    return TIPOS_DOCUMENTOS.get(tipo_chave, {}).get('botao_pdf', 'Outros')

def obter_padroes_tipo_documento(tipo_chave: str) -> List[str]:
    """Obtém os padrões de identificação para um tipo de documento."""
    return TIPOS_DOCUMENTOS.get(tipo_chave, {}).get('padroes', [])

def listar_tipos_documento() -> List[str]:
    """Lista todas as chaves de tipos de documento disponíveis."""
    return list(TIPOS_DOCUMENTOS.keys())


# ================================
# FUNÇÕES DE NORMALIZAÇÃO
# ================================

def normalizar(s: Union[str, Any]) -> Union[str, Any]:
    """Normaliza string removendo acentos e convertendo para lowercase."""
    if isinstance(s, str):
        s = s.strip().lower()
        s = ''.join(
            c for c in unicodedata.normalize('NFD', s)
            if unicodedata.category(c) != 'Mn'
        )
    return s


def _gerar_regex_palavra_chave(palavra_chave: str) -> Optional[str]:
    """Gera regex flexivel para busca de palavra-chave normalizada."""
    palavra_norm = normalizar(palavra_chave)
    if not isinstance(palavra_norm, str) or not palavra_norm:
        return None

    separador_flex = r'[^a-z0-9]*'
    partes = []
    tamanho = len(palavra_norm)

    for i, caractere in enumerate(palavra_norm):
        proximo = palavra_norm[i + 1] if i + 1 < tamanho else ''

        if caractere.isalnum():
            partes.append(re.escape(caractere))
        else:
            if not partes or partes[-1] != separador_flex:
                partes.append(separador_flex)
            continue

        if proximo and ((caractere.isalpha() and proximo.isdigit()) or (caractere.isdigit() and proximo.isalpha())):
            partes.append(separador_flex)

    padrao = ''.join(partes)
    return rf'(?<![a-z0-9]){padrao}(?![a-z0-9])'


def contar_ocorrencias_palavra_chave(texto: str, palavra_chave: str) -> int:
    """
    Conta ocorrencias de palavra-chave com padrao flexivel para separadores.

    Exemplo: "ipv6" casa com "ipv6", "ipv-6" e "ipv 6".
    """
    if not texto or not palavra_chave:
        return 0

    texto_norm = normalizar(texto)
    if not isinstance(texto_norm, str):
        return 0

    regex = _gerar_regex_palavra_chave(palavra_chave)
    if not regex:
        return 0

    return len(re.findall(regex, texto_norm, flags=re.IGNORECASE))


def extrair_contextos_palavra_chave(texto: str, palavra_chave: str, limite: int = 3, janela: int = 45) -> List[str]:
    """Extrai pequenos trechos onde a palavra-chave foi identificada."""
    if not texto or not palavra_chave:
        return []

    texto_norm = normalizar(texto)
    if not isinstance(texto_norm, str):
        return []

    regex = _gerar_regex_palavra_chave(palavra_chave)
    if not regex:
        return []

    contextos = []
    for match in re.finditer(regex, texto_norm, flags=re.IGNORECASE):
        ini = max(0, match.start() - janela)
        fim = min(len(texto_norm), match.end() + janela)
        trecho = texto_norm[ini:fim].replace('\n', ' ')
        trecho = re.sub(r'\s+', ' ', trecho).strip()
        if trecho:
            contextos.append(trecho)
        if len(contextos) >= limite:
            break

    return contextos


def normalizar_dados(dados: Dict[str, Any]) -> Dict[str, Any]:
    """Normaliza todos os dados string em um dicionário."""
    for k, v in dados.items():
        if isinstance(v, list):
            dados[k] = [normalizar(item) if isinstance(item, str) else item for item in v]
        elif isinstance(v, str):
            dados[k] = normalizar(v)
    return dados


# ================================
# FUNÇÕES DE BUSCA EM JSON
# ================================

def buscar_valor(json_data: Union[List[Dict], Dict], key_busca: str, valor_busca: Any, key_retorno: str) -> Optional[Any]:
    """
    Busca em uma lista de dicionários (ou JSON similar)
    onde key_busca == valor_busca e retorna o valor de key_retorno.
    """
    if isinstance(json_data, list):
        for item in json_data:
            if isinstance(item, dict):
                if item.get(key_busca) == valor_busca:
                    return item.get(key_retorno)
    elif isinstance(json_data, dict):
        if json_data.get(key_busca) == valor_busca:
            return json_data.get(key_retorno)
    return None


# ================================
# FUNÇÕES DE VERSIONAMENTO
# ================================

def obter_versao_git() -> str:
    """Obtém a versão do script baseada em git tag ou commit hash."""
    try:
        # Tentar obter a tag mais recente (ordenada por versão)
        resultado = subprocess.run(
            GIT_COMMANDS['tags'],
            capture_output=True,
            text=True,
            timeout=GIT_TIMEOUT
        )
        
        if resultado.returncode == 0 and resultado.stdout.strip():
            # Pegar a primeira linha (tag mais recente)
            tags = resultado.stdout.strip().split('\n')
            if tags and tags[0]:
                return tags[0]
        
        # Fallback: tentar describe --tags
        resultado = subprocess.run(
            GIT_COMMANDS['describe'],
            capture_output=True,
            text=True,
            timeout=GIT_TIMEOUT
        )
        
        if resultado.returncode == 0 and resultado.stdout.strip():
            return resultado.stdout.strip()
        
        # Se não há tags, tentar obter o hash do commit
        resultado = subprocess.run(
            GIT_COMMANDS['commit_hash'],
            capture_output=True,
            text=True,
            timeout=GIT_TIMEOUT
        )
        
        if resultado.returncode == 0 and resultado.stdout.strip():
            return f"commit-{resultado.stdout.strip()}"
            
    except (subprocess.TimeoutExpired, FileNotFoundError, Exception):
        pass
    
    # Fallback para versão baseada em data
    return VERSAO_PADRAO.format(datetime.now().strftime('%Y.%m.%d'))


# ================================
# FUNÇÕES DE SISTEMA E PATHS
# ================================

def is_bundled() -> bool:
    """Verifica se o código está rodando como executável PyInstaller."""
    return getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS')


def get_files_folder() -> str:
    """Retorna o diretório base para arquivos baseado no modo de execução."""
    if is_bundled():
        folder = os.path.dirname(sys.executable)
        #log_info(MENSAGENS_STATUS['modo_executavel'].format(folder))
        return folder
    else:
        #log_info(MENSAGENS_STATUS['modo_script'].format(TBN_FILES_FOLDER))
        return TBN_FILES_FOLDER


def get_profile_dir() -> str:
    """Retorna o diretório do perfil Chrome baseado no modo de execução."""
    if is_bundled():
        files_folder = os.path.dirname(sys.executable)
        return os.path.join(files_folder, CHROME_PROFILE_DIR)
    else:
        return os.path.join(Path(__file__).parent.parent, CHROME_PROFILE_DIR)


def req_para_fullpath(req: str) -> str:
    """Converte número do requerimento (num/ano) para caminho completo da pasta"""
    num, ano = req.split("/")
    requerimento = rf"{REQUERIMENTOS_DIR_INBOX}\_{ano}.{num}"
    full_path = os.path.join(get_files_folder(), requerimento)
    return full_path

def req_para_usedpath(req: str) -> str:
    """Converte número do requerimento (num/ano) para caminho completo da pasta"""
    num, ano = req.split("/")
    requerimento = rf"{REQUERIMENTOS_DIR_INBOX}\{ano}.{num}"
    full_path = os.path.join(get_files_folder(), requerimento)
    return full_path

def fullpath_para_req(nome_diretorio: str) -> str:
    """
    Converte nome do diretório (formato YY.num ou _YY.num) para número do requerimento (num/ano)
    
    Args:
        nome_diretorio (str): Nome do diretório no formato "YY.12345" ou "_YY.12345"
    
    Returns:
        str: Requerimento no formato "12345/YYYY"
    
    Examples:
        >>> fullpath_para_req("25.12345")
        "12345/25"
        >>> fullpath_para_req("_25.12345")
        "12345/25"
        >>> fullpath_para_req("24.98765")
        "98765/24"
    """
    # Remove underscore inicial se existir
    nome_limpo = nome_diretorio.lstrip('_')
    
    # Divide por ponto para separar ano e número
    partes = nome_limpo.split('.')
    
    if len(partes) != 2:
        raise ValueError(f"Formato inválido: {nome_diretorio}. Esperado: 'YY.numero' ou '_YY.numero'")
    
    ano_curto, numero = partes
    
    # Converte ano de 2 dígitos para 4 dígitos
    # Assume que anos 00-99 são 2000-2099
    ano_completo = f"{ano_curto.zfill(2)}"
    
    return f"{numero}/{ano_completo}"


def criar_pasta_se_nao_existir(req: str) -> str:
    """Cria pasta do requerimento se não existir e retorna o caminho"""
    full_path = req_para_fullpath(req)
    used_path = req_para_usedpath(req)
    if os.path.exists(used_path):
        # Se used_path existe, renomear para adicionar underscore
        # Extrai o nome do último diretório e adiciona _
        parent_dir = os.path.dirname(used_path)
        dir_name = os.path.basename(used_path)
        new_dir_name = f"_{dir_name}"
        new_path = os.path.join(parent_dir, new_dir_name)
        
        # Verificar se o diretório de destino já existe
        if not os.path.exists(new_path):
            try:
                os.rename(used_path, new_path)
                log_info(f"📁 Diretório renomeado: {used_path} -> {new_path}")
            except OSError as e:
                log_erro(f"Erro ao renomear diretório {used_path}: {e}")
        else:
            log_info(f"📁 Diretório com underscore já existe, removendo duplicata: {used_path}")
            try:
                # Remove o diretório sem underscore se o com underscore já existe
                import shutil
                shutil.rmtree(used_path)
                log_info(f"📁 Diretório duplicata removido: {used_path}")
            except OSError as e:
                log_erro(f"Erro ao remover diretório duplicata {used_path}: {e}")
    elif not os.path.exists(full_path):
        os.makedirs(full_path, exist_ok=True)
        log_info(f"📁 Pasta criada: {full_path}")
    return full_path


# ================================
# FUNÇÕES DE MANIPULAÇÃO DE ARQUIVOS
# ================================

def carregar_json(caminho: Union[str, Path], encoding: str = 'utf-8') -> Optional[Union[Dict, List]]:
    """
    Carrega dados de um arquivo JSON de forma segura.
    
    Args:
        caminho: Caminho para o arquivo JSON
        encoding: Codificação do arquivo (padrão: utf-8)
    
    Returns:
        Dados do JSON ou None em caso de erro
    """
    try:
        with open(caminho, 'r', encoding=encoding) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError, Exception):
        return None


def carregar_json_com_fallback(caminho: str) -> Dict:
    """
    Carrega arquivo JSON de configuração com logging de erro e fallback para dict vazio.
    Função migrada do AnalisadorRequerimentos para uso geral no projeto.
    
    Args:
        caminho: Caminho para o arquivo JSON de configuração
        
    Returns:
        Dict com os dados do JSON ou dict vazio em caso de erro
    """
    try:
        with open(caminho, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        log_erro(f"Arquivo de configuração não encontrado: {caminho}")
        return {}
    except json.JSONDecodeError:
        log_erro(f"Erro ao decodificar JSON: {caminho}")
        return {}
    except Exception as e:
        log_erro(f"Erro inesperado ao carregar JSON {caminho}: {str(e)}")
        return {}


def salvar_json(dados: Union[Dict, List], caminho: Union[str, Path], encoding: str = 'utf-8', indent: int = 2) -> bool:
    """
    Salva dados em um arquivo JSON de forma segura.
    
    Args:
        dados: Dados para salvar
        caminho: Caminho para o arquivo JSON
        encoding: Codificação do arquivo (padrão: utf-8)
        indent: Indentação para formatação (padrão: 2)
    
    Returns:
        True se salvou com sucesso, False caso contrário
    """
    try:
        with open(caminho, 'w', encoding=encoding) as f:
            json.dump(dados, f, ensure_ascii=False, indent=indent)
        return True
    except Exception:
        return False


# ================================
# FUNÇÕES DE VALIDAÇÃO
# ================================

def validar_cnpj(cnpj: str) -> bool:
    """
    Valida se um CNPJ possui formato válido (apenas estrutura, não algoritmo).
    
    Args:
        cnpj: CNPJ para validar (formatado ou não)
    
    Returns:
        True se o formato é válido, False caso contrário
    """
    if not cnpj:
        return False
    
    # Remove formatação
    cnpj_numeros = desformatar_cnpj(cnpj)
    
    # Verifica se tem 14 dígitos
    return len(cnpj_numeros) == 14 and cnpj_numeros.isdigit()


def validar_caminho_arquivo(caminho: Union[str, Path]) -> bool:
    """
    Valida se um caminho de arquivo existe e é um arquivo válido.
    
    Args:
        caminho: Caminho para validar
    
    Returns:
        True se é um arquivo válido, False caso contrário
    """
    try:
        path_obj = Path(caminho)
        return path_obj.exists() and path_obj.is_file()
    except Exception:
        return False


def extrair_normas_por_padrao(content: str) -> List[str]:
    """
    Extrai normas usando padrões específicos (função utilitária centralizada).
    Baseado no método _extract_normas_by_pattern como modelo integral.
    """
    normas = []
    custom_patterns = ['ATO', 'RESOLUÇÃO']
    
    normas_section = limpar_texto(content, palavras=["contato","Contato","Nº","N°","NO","nº","n°","n.","N.","no","de","do", "da", "anatel"], simbolos=["."])

    lines = [
        subcampo.strip()
        for linha in normas_section.split('\n')
        for campo in linha.split(';')
        for subcampo in campo.split(',')
    ]
    for line in lines:
        line = line.strip()
        for pattern in custom_patterns:
            if pattern in normas_section.upper():
                # Regex corrigida - o problema era \s+ que exige pelo menos 1 espaço
                # Mudei para \s* para permitir zero ou mais espaços
                norma_matches = re.findall(
                    r'(ATO|RESOLUÇÃO|RESOLUÇÕES?)\s*(?:\([^)]+\))?\s*(?:da\s+\w+\s+)?(?:|Nº|N°|NO|nº|n°|no|n.|N.)?[\s:]*(\d+)',
                    normas_section,
                    re.IGNORECASE
                )                       
                
                # Processar cada match
                for tipo, numero in norma_matches:
                    tipo_normalizado = tipo.lower()
                    if 'resolu' in tipo_normalizado:
                        tipo_normalizado = 'resolucao'
                    else:
                        tipo_normalizado = 'ato'
                    
                    norma_formatada = f"{tipo_normalizado}{numero}"
                    if norma_formatada not in normas:
                        normas.append(norma_formatada)
                
                break

    return normas


def validar_caminho_diretorio(caminho: Union[str, Path]) -> bool:
    """
    Valida se um caminho de diretório existe e é um diretório válido.
    
    Args:
        caminho: Caminho para validar
    
    Returns:
        True se é um diretório válido, False caso contrário
    """
    try:
        path_obj = Path(caminho)
        return path_obj.exists() and path_obj.is_dir()
    except Exception:
        return False


# ================================
# FUNÇÕES DE REQUERIMENTOS E EXCEL
# ================================

def processar_requerimentos_excel(num_req: str) -> None:
    """
    Processa requerimentos para atualização da planilha Excel ORCN.
    
    Lê arquivos JSON de requerimentos e atualiza a planilha Excel se o número
    do requerimento não existir na coluna B da aba "Requerimentos-Análise".
    
    Args:
        num_req (str): Número do requerimento no formato "xx.xxxxx" ou "*" para todos
    
    Raises:
        ImportError: Se pandas ou openpyxl não estiverem disponíveis
        FileNotFoundError: Se arquivos necessários não forem encontrados
        ValueError: Se formato do requerimento for inválido
    """    
    
    
    # Validar se arquivos e diretórios existem
    if not os.path.exists(EXCEL_PATH):
        log_info(f"Erro: Planilha não encontrada: {EXCEL_PATH}")
        return
    
    if not os.path.exists(REQUERIMENTOS_PATH):
        log_info(f"Erro: Diretório de requerimentos não encontrado: {REQUERIMENTOS_PATH}")
        return
    
    try:
        # Carregar planilha Excel
        df = pd.read_excel(EXCEL_PATH, sheet_name=EXCEL_SHEET_NAME)
        #log_info(f"Planilha carregada com {len(df)} linhas")
        
        # Obter lista de requerimentos já existentes na coluna B
        coluna_req = df.columns[TAB_REQUERIMENTOS['num_req']]  # Coluna B (índice 1)
        requerimentos_existentes = set(df[coluna_req].dropna().astype(str))
        #log_info(f"Requerimentos existentes na planilha: {len(requerimentos_existentes)}")
        
        # Determinar quais requerimentos processar
        if num_req == "*":
            # Processar todos os requerimentos do diretório
            #log_info("Processando todos os requerimentos do diretório...")
            requerimentos_para_processar = []
            
            for item in os.listdir(REQUERIMENTOS_PATH):
                caminho_item = os.path.join(REQUERIMENTOS_PATH, item)
                if os.path.isdir(caminho_item) and re.match(r'^\d{2}\.\d{5}$', item):
                    requerimentos_para_processar.append(item)
            
            #log_info(f"Encontrados {len(requerimentos_para_processar)} diretórios de requerimentos")
        else:
            # Processar requerimento específico
            if not re.match(r'^\d{2}\.\d{5}$', num_req):
                log_info(f"Erro: Formato inválido do requerimento: {num_req}. Use formato xx.xxxxx")
                return
            
            requerimentos_para_processar = [num_req]
            #log_info(f"Processando requerimento específico: {num_req}")
        
        # Processar cada requerimento
        requerimentos_adicionados = []
        requerimentos_ja_existentes = []
        requerimentos_com_erro = []
        novas_linhas_para_excel = []  # Armazenar linhas para adicionar diretamente no Excel
        
        for req in requerimentos_para_processar:
            try:
                # Verificar se já existe na planilha (formato pode variar)
                formatos_possíveis = [
                    req,  # formato 25.06969
                    req.replace('.', '/'),  # formato 25/06969 (como aparece no JSON)
                    f"{req.split('.')[1]}/{req.split('.')[0]}"  # formato 06969/25
                ]
                
                req_existe = any(fmt in requerimentos_existentes for fmt in formatos_possíveis)
                
                if req_existe:
                    log_info(f"Requerimento {req} já existe na planilha")
                    requerimentos_ja_existentes.append(req)
                    continue
                
                # Ler arquivo JSON do requerimento
                pasta_req = "_" + req
                arquivo_json = os.path.join(REQUERIMENTOS_PATH, pasta_req, f"{req}.json")
                
                if not os.path.exists(arquivo_json):
                    log_info(f"Aviso: Arquivo JSON não encontrado: {arquivo_json}")
                    requerimentos_com_erro.append(req)
                    continue
                
                # Carregar dados do JSON
                with open(arquivo_json, 'r', encoding='utf-8') as f:
                    dados_req = json.load(f)
                
                # Extrair informações do JSON conforme estrutura esperada
                req_data = dados_req.get('requerimento', {})
                
                # Mapear dados do JSON para colunas da planilha
                nova_linha = _mapear_dados_json_para_excel(req_data)
                
                if nova_linha:
                    # Armazenar linha para adicionar no Excel
                    novas_linhas_para_excel.append(nova_linha)
                    requerimentos_adicionados.append(req)
                    #log_info(f"Requerimento {req} preparado para adição à planilha")
                else:
                    log_info(f"Erro: Não foi possível mapear dados do requerimento {req}")
                    requerimentos_com_erro.append(req)
                    
            except Exception as e:
                log_info(f"Erro ao processar requerimento {req}: {e}")
                requerimentos_com_erro.append(req)
        
        # Salvar planilha se houver alterações
        if requerimentos_adicionados and novas_linhas_para_excel:
            # Usar openpyxl para preservar todas as abas existentes
            wb = load_workbook(EXCEL_PATH)
            ws = wb[EXCEL_SHEET_NAME]
            
            # Encontrar a próxima linha vazia
            proxima_linha = ws.max_row + 1
            
            # Adicionar cada nova linha diretamente no Excel
            for nova_linha in novas_linhas_para_excel:
                for col_idx, valor in enumerate(nova_linha):
                    ws.cell(row=proxima_linha, column=col_idx + 1, value=valor)
                proxima_linha += 1
            
            # Salvar preservando todas as abas
            wb.save(EXCEL_PATH)
            wb.close()
            #log_info(f"Planilha atualizada com {len(requerimentos_adicionados)} novos requerimentos (todas as abas preservadas)")
        else:
            log_info("Nenhum requerimento novo para adicionar")
        
        # Relatório final
        #log_info("\n=== RELATÓRIO DE PROCESSAMENTO ===")
        #log_info(f"Requerimentos processados: {len(requerimentos_para_processar)}")
        #log_info(f"Requerimentos adicionados: {len(requerimentos_adicionados)}")
        #log_info(f"Requerimentos já existentes: {len(requerimentos_ja_existentes)}")
        #log_info(f"Requerimentos com erro: {len(requerimentos_com_erro)}")        
        #if requerimentos_adicionados:
        #    log_info(f"Novos requerimentos: {', '.join(requerimentos_adicionados)}")        
        #if requerimentos_com_erro:
        #    log_info(f"Requerimentos com erro: {', '.join(requerimentos_com_erro)}")
            
    except Exception as e:
        log_info(f"Erro ao processar planilha: {e}")


def _converter_para_excel(valor: Any) -> Any:
    """
    Converte valores para formatos compatíveis com Excel.
    
    Args:
        valor: Valor a ser convertido
    
    Returns:
        Valor convertido para formato compatível com Excel
    """
    if valor is None:
        return None
    elif isinstance(valor, list):
        # Converter lista para string separada por vírgulas
        return ', '.join(str(item) for item in valor)
    elif isinstance(valor, dict):
        # Converter dicionário para string representativa
        return str(valor)
    elif isinstance(valor, (int, float, bool)):
        return valor
    else:
        # Para strings e outros tipos, converter para string
        return str(valor)


def _mapear_dados_json_para_excel(req_data: Dict[str, Any]) -> Optional[List[Any]]:
    """
    Mapeia dados do JSON do requerimento para formato da planilha Excel.
    
    Args:
        req_data: Dados do requerimento extraídos do JSON
    
    Returns:
        Lista com dados mapeados para colunas da planilha ou None se erro
    """
    try:
        
        
        # Criar linha vazia com tamanho correto (11 colunas total)
        nova_linha = [None] * 11
        
        # Mapear campos do JSON para posições da planilha
        # Coluna A (índice 0): 'Análise' - deixar vazio ou colocar valor padrão
        #nova_linha[0] = 'AUTOMATICO'
        
        # Coluna B (índice 1): 'Nº do Requerimento'
        nova_linha[TAB_REQUERIMENTOS['num_req']] = _converter_para_excel(req_data.get('num_req', ''))
        
        # Coluna C (índice 2): 'Nº de Homologação'
        nova_linha[TAB_REQUERIMENTOS['cod_homologacao']] = _converter_para_excel(req_data.get('cod_homologacao', ''))
        
        # Coluna D (índice 3): 'Nº do Certificado'
        nova_linha[TAB_REQUERIMENTOS['num_cct']] = _converter_para_excel(req_data.get('num_cct', ''))
        
        # Coluna E (índice 4): 'Tipo do Produto'
        nova_linha[TAB_REQUERIMENTOS['tipo_equipamento']] = _converter_para_excel(req_data.get('tipo_equipamento', ''))
        
        # Coluna F (índice 5): 'Modelo'
        nova_linha[TAB_REQUERIMENTOS['modelos']] = _converter_para_excel(req_data.get('modelos', ''))
        
        # Coluna G (índice 6): 'Solicitante'
        nova_linha[TAB_REQUERIMENTOS['solicitante']] = _converter_para_excel(req_data.get('solicitante', ''))
        
        # Coluna H (índice 7): 'Fabricante'
        nova_linha[TAB_REQUERIMENTOS['fabricante']] = _converter_para_excel(req_data.get('fabricante', ''))
        
        # Coluna I (índice 8): 'Data da Conclusão'
        data = req_data.get('data', '')
        if data:
            # Converter formato de data se necessário
            try:
                # Assumindo formato DD/MM/YYYY do JSON
                data_obj = datetime.strptime(data, '%d/%m/%Y')
                nova_linha[TAB_REQUERIMENTOS['data']] = _converter_para_excel(data_obj)
            except ValueError:
                # Manter string original se conversão falhar
                nova_linha[TAB_REQUERIMENTOS['data']] = _converter_para_excel(data)
        
        # Coluna J (índice 9): 'Situação'
        nova_linha[TAB_REQUERIMENTOS['status']] = _converter_para_excel(req_data.get('status', 'Em Análise'))
        
        # Coluna K (índice 10): 'Tempo' - deixar vazio
        nova_linha[10] = None
        
        return nova_linha
        
    except Exception as e:
        log_info(f"Erro ao mapear dados JSON: {e}")
        return None


# ================================
# FUNÇÕES DE CONTROLE DE LOG DE DOWNLOADS
# ================================

def get_download_log_path() -> str:
    """Retorna o caminho completo para o arquivo de log de downloads."""
    files_folder = get_files_folder()
    return os.path.join(files_folder, DOWNLOAD_LOG_FILENAME)


def carregar_log_downloads() -> Dict[str, Dict]:
    """
    Carrega o log de status dos downloads dos requerimentos.
    
    Returns:
        Dict[str, Dict]: Dicionário com status dos downloads por requerimento.
                        Formato: {
                            "requerimento": {
                                "status": "completed" | "in_progress" | "failed",
                                "timestamp": "2025-01-01 10:30:00",
                                "arquivos_baixados": 5,
                                "erro": "mensagem de erro (se houver)"
                            }
                        }
    """
    log_path = get_download_log_path()
    log_data = carregar_json(log_path)
    # Garantir que sempre retorna um dicionário
    if isinstance(log_data, dict):
        return log_data
    return {}


def salvar_log_downloads(log_data: Dict[str, Dict]) -> bool:
    """
    Salva o log de status dos downloads dos requerimentos.
    
    Args:
        log_data: Dicionário com status dos downloads por requerimento
        
    Returns:
        bool: True se salvou com sucesso, False caso contrário
    """
    log_path = get_download_log_path()
    return salvar_json(log_data, log_path, indent=2)


def requerimento_ja_baixado(requerimento: str) -> bool:
    """
    Verifica se um requerimento já foi baixado com sucesso.
    
    Args:
        requerimento: Número do requerimento no formato "num/ano"
        
    Returns:
        bool: True se o requerimento já foi baixado completamente
    """
    log_data = carregar_log_downloads()
    req_status = log_data.get(requerimento, {})
    return req_status.get("status") == "completed"


def marcar_requerimento_em_progresso(requerimento: str) -> bool:
    """
    Marca um requerimento como em progresso no log de downloads.
    
    Args:
        requerimento: Número do requerimento no formato "num/ano"
        
    Returns:
        bool: True se marcou com sucesso
    """
    log_data = carregar_log_downloads()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    log_data[requerimento] = {
        "status": "in_progress",
        "timestamp": timestamp,
        "arquivos_baixados": 0
    }
    
    return salvar_log_downloads(log_data)


def marcar_requerimento_concluido(requerimento: str, arquivos_baixados: int) -> bool:
    """
    Marca um requerimento como concluído no log de downloads.
    
    Args:
        requerimento: Número do requerimento no formato "num/ano"
        arquivos_baixados: Quantidade de arquivos baixados
        
    Returns:
        bool: True se marcou com sucesso
    """
    log_data = carregar_log_downloads()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    log_data[requerimento] = {
        "status": "completed",
        "timestamp": timestamp,
        "arquivos_baixados": arquivos_baixados
    }
    
    return salvar_log_downloads(log_data)


def marcar_requerimento_com_erro(requerimento: str, erro: str) -> bool:
    """
    Marca um requerimento com erro no log de downloads.
    
    Args:
        requerimento: Número do requerimento no formato "num/ano"
        erro: Mensagem de erro
        
    Returns:
        bool: True se marcou com sucesso
    """
    log_data = carregar_log_downloads()
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    log_data[requerimento] = {
        "status": "failed",
        "timestamp": timestamp,
        "erro": erro,
        "arquivos_baixados": 0
    }
    
    return salvar_log_downloads(log_data)


def limpar_log_downloads_se_completo(requerimentos_processados: List[str]) -> bool:
    """
    Limpa o log de downloads se todos os requerimentos da lista foram concluídos.
    
    Args:
        requerimentos_processados: Lista de requerimentos que foram processados na sessão atual
        
    Returns:
        bool: True se o log foi limpo (todos concluídos), False caso contrário
    """
    log_data = carregar_log_downloads()
    
    # Verifica se todos os requerimentos processados foram concluídos
    todos_concluidos = True
    for req in requerimentos_processados:
        req_status = log_data.get(req, {})
        if req_status.get("status") != "completed":
            todos_concluidos = False
            break
    
    if todos_concluidos and requerimentos_processados:
        # Limpa o log
        log_path = get_download_log_path()
        try:
            if os.path.exists(log_path):
                #os.remove(log_path)
                log_info("🗑️ Log de downloads limpo - todos os requerimentos foram processados com sucesso")
            return True
        except Exception as e:
            log_erro(f"Erro ao limpar log de downloads: {e}")
            return False
    
    return False


def obter_requerimentos_pendentes(todos_requerimentos: List[str]) -> List[str]:
    """
    Filtra lista de requerimentos removendo os que já foram baixados com sucesso.
    
    Args:
        todos_requerimentos: Lista com todos os requerimentos encontrados
        
    Returns:
        List[str]: Lista apenas com requerimentos que ainda precisam ser baixados
    """
    requerimentos_pendentes = []
    requerimentos_ja_baixados = []
    
    for req in todos_requerimentos:
        if requerimento_ja_baixado(req):
            requerimentos_ja_baixados.append(req)
        else:
            requerimentos_pendentes.append(req)
    
    if requerimentos_ja_baixados:
        log_info(f"✅ {len(requerimentos_ja_baixados)} requerimento(s) já baixado(s): {', '.join(requerimentos_ja_baixados)}")
    
    if requerimentos_pendentes:
        log_info(f"⏳ {len(requerimentos_pendentes)} requerimento(s) pendente(s): {', '.join(requerimentos_pendentes)}")
    else:
        log_info("✅ Todos os requerimentos já foram baixados!")
    
    return requerimentos_pendentes


def extract_pdf_content_from_ocr(pdf_path: Path) -> Optional[str]:
    """
    Extrai conteúdo de PDF usando OCR (Tesseract).
    Função utilitária reutilizável em todo o projeto.
    
    Args:
        pdf_path: Path para o arquivo PDF
        
    Returns:
        str com o texto extraído ou None em caso de erro
    """
    if not OCR_DISPONIVEL:
        log_erro("Dependências de OCR não disponíveis (pdf2image, pytesseract)")
        return None
    
    try:
        # Configurar caminho do Tesseract se necessário
        try:
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_PATH  # type: ignore
        except:
            pass
        
        # Converte cada página do PDF em imagem
        paginas = convert_from_path(pdf_path)  # type: ignore
        texto_completo = ""

        # Extrai texto de cada página via OCR
        for i, pagina in enumerate(paginas, start=1):
            texto_pagina = pytesseract.image_to_string(pagina, lang='por')  # type: ignore
            texto_completo += texto_pagina
            
        return texto_completo
    
    except Exception as e:
        log_erro(f"Falha ao extrair por OCR {pdf_path.name}: {e}")
        return None


def testar_radiacao_restrita(nome_equipamento: str) -> bool:
    """
    Testa se um equipamento é do tipo "Radiação Restrita" buscando no arquivo equipamentos.json.
    
    Busca pelo nome exato do equipamento e verifica se o ID correspondente é "EQ078".
    
    Args:
        nome_equipamento (str): Nome do equipamento para buscar
        
    Returns:
        bool: True se o equipamento for de radiação restrita (EQ078), False caso contrário
    """
    try:
        # Carregar dados do arquivo equipamentos.json        
        equipamentos = carregar_json_com_fallback(JSON_FILES['equipamentos'])
        
        # Buscar equipamento por nome exato usando a função utilitária
        id_equipamento = buscar_valor(equipamentos, 'nome', nome_equipamento, 'id')
        
        # Verificar se o ID encontrado é EQ093 (Radiação Restrita)
        return id_equipamento in ['EQ093', 'EQ088', 'EQ078', 'EQ053'] 
        
    except Exception as e:
        log_erro(f"Erro ao testar radiação restrita para '{nome_equipamento}': {e}")
        return False