from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
import os
import re
import time

# -------- Configurações --------
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
FILES_FOLDER = r"C:\Users\tbnobrega\OneDrive - ANATEL\Anatel\_ORCN"
PROFILE_DIR = os.path.abspath("./meu_perfil_chrome")
LOG_REQUERIMENTOS = os.path.join(FILES_FOLDER, 'ORCN - Documentação pertinente.xlsx')
LISTA_URL = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"

# -------- Funções --------
def criar_pasta_se_nao_existir(req):
    num, ano = req.split("/")
    requerimento = rf"Requerimentos\{ano}.{num}"
    full_path = os.path.join(FILES_FOLDER, requerimento)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
    return full_path

def atualizar_excel(rows):
    novosAnalises = []
    indicesNovos = []
    wb = load_workbook(LOG_REQUERIMENTOS)
    ws = wb['Requerimentos-Análise']
    tabRequerimentos = ws.tables['tabRequerimentos']

    linhas_existentes = [tuple(r) for r in ws.iter_rows(values_only=True)]

    for i, row in enumerate(rows, start=1):
        cols = row.find_elements(By.TAG_NAME, "td")
        dados = [col.text.strip() for col in cols]
        
        # Verificar se dados tem elementos suficientes
        if len(dados) < 10:
            print(f"⚠ Linha {i} com dados insuficientes: {len(dados)} colunas")
            continue
            
        dados[0] = 'AUTOMATICO'

        if dados[9] == 'Em Análise':
            novalinha = dados[:10]
            if tuple(novalinha) not in linhas_existentes:
                ws.append(novalinha)
                print(f"Linha adicionada: {novalinha}")
                novosAnalises.append(novalinha[1])
                indicesNovos.append(i)
                criar_pasta_se_nao_existir(novalinha[1])
            else:
                print("Linha já existe, não adicionada.")

    # Atualiza o range da tabela
    ultima_linha = ws.max_row
    col_inicio, linha_inicio = tabRequerimentos.ref.split(":")[0][0], tabRequerimentos.ref.split(":")[0][1:]
    col_fim = tabRequerimentos.ref.split(":")[1][0]
    tabRequerimentos.ref = f"{col_inicio}{linha_inicio}:{col_fim}{ultima_linha}"

    wb.save(LOG_REQUERIMENTOS)
    wb.close()
    return indicesNovos, novosAnalises

def salvarPDFs():
    # Aqui você coloca a lógica de salvar PDFs da aba aberta
    print("📄 Salvando PDFs... (função ainda precisa ser implementada)")

def encontrar_botao_visualizar(row):
    """Tenta múltiplos seletores para encontrar o botão de visualizar"""
    seletores = [
        "button[title='Visualizar em Tela cheia']",
        "button[title*='Visualizar']",
        "button[title*='Tela cheia']",
        "button.ui-button",
        "button[onclick*='window.open']",
        ".//button[@title='Visualizar em Tela cheia']",
        ".//button[contains(@title, 'Visualizar')]",
        ".//button[contains(@onclick, 'window.open')]"
    ]
    
    for seletor in seletores:
        try:
            if seletor.startswith(".//"):
                # Usar XPath
                btn = row.find_element(By.XPATH, seletor)
            else:
                # Usar CSS Selector
                btn = row.find_element(By.CSS_SELECTOR, seletor)
            
            if btn:
                return btn
        except NoSuchElementException:
            continue
    
    # Se não encontrou, procura todos os botões na linha
    try:
        botoes = row.find_elements(By.TAG_NAME, "button")
        print(f"   🔍 Encontrados {len(botoes)} botões na linha")
        for idx, botao in enumerate(botoes):
            title = botao.get_attribute("title")
            onclick = botao.get_attribute("onclick")
            print(f"      Botão {idx}: title='{title}', onclick presente={onclick is not None}")
            
            # Procura por botão com onclick que abre janela
            if onclick and "window.open" in onclick:
                return botao
    except Exception as e:
        print(f"   ⚠ Erro ao procurar botões: {e}")
    
    return None

# -------- Script principal --------
# Configurar opções do Chrome
chrome_options = Options()
chrome_options.binary_location = CHROME_PATH
chrome_options.add_argument(f"--user-data-dir={PROFILE_DIR}")
chrome_options.add_argument("--start-maximized")
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])

# Inicializar o driver
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 30)

try:
    # Navegar para a URL
    print("🌐 Acessando a página...")
    driver.get(LISTA_URL)
    
    # Aguardar o carregamento completo da página
    time.sleep(3)
    
    # Tentar diferentes seletores para o botão
    print("🔍 Procurando botão 'todos'...")
    btn_todos = None
    
    seletores = [
        "#menuForm\\:todos",
        "[id='menuForm:todos']",
        "button[id='menuForm:todos']",
    ]
    
    for seletor in seletores:
        try:
            btn_todos = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, seletor)))
            print(f"✅ Botão encontrado com seletor: {seletor}")
            
            wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, seletor)))
            driver.execute_script("arguments[0].scrollIntoView(true);", btn_todos)
            time.sleep(0.5)
            
            try:
                btn_todos.click()
            except:
                driver.execute_script("arguments[0].click();", btn_todos)
            
            print("✅ Botão clicado!")
            break
        except Exception as e:
            print(f"⚠ Seletor {seletor} não funcionou")
            continue
    
    if btn_todos is None:
        raise Exception("Botão 'todos' não encontrado")
    
    # Aguardar carregamento da tabela após o clique
    print("⏳ Aguardando carregamento da tabela...")
    time.sleep(3)
    
    # Tentar diferentes seletores para a tabela
    seletores_tabela = [
        "#form\\:tarefasTable_data tr",
        "[id='form:tarefasTable_data'] tr",
        "tbody[id='form:tarefasTable_data'] tr"
    ]
    
    rows = None
    for seletor in seletores_tabela:
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, seletor)))
            rows = driver.find_elements(By.CSS_SELECTOR, seletor)
            print(f"✅ Tabela encontrada com seletor: {seletor}")
            break
        except:
            continue
    
    if rows is None or len(rows) == 0:
        raise Exception("Tabela não encontrada ou vazia")
    
    print(f"🔎 {len(rows)} processos na caixa de entrada")

    # Atualiza Excel e cria pastas
    novos_indices, novas_analises = atualizar_excel(rows)
    print("✅ Excel atualizado")
    
    if not novos_indices:
        print("ℹ️ Nenhum novo processo para abrir")
    else:
        print(f"📋 {len(novos_indices)} novos processos para abrir")

    # Armazena a janela original
    janela_original = driver.current_window_handle

    # Processa cada linha nova
    for i, row in enumerate(rows, start=1):
        if i not in novos_indices:
            continue

        print(f"\n▶ Processando linha {i}...")
        
        # Encontrar o botão de visualizar
        btn = encontrar_botao_visualizar(row)
        
        if btn is None:
            print(f"⚠ Botão não encontrado na linha {i}")
            # Debug: mostrar HTML da linha
            try:
                html_snippet = row.get_attribute('innerHTML')
                if html_snippet:
                    print(f"   HTML da linha: {html_snippet[:200]}...")
            except:
                pass
            continue

        btn.Click()

        input("A página foi aberta? Pressione Enter para continuar...")
        

except Exception as e:
    print(f"❌ Erro: {str(e)}")
    import traceback
    traceback.print_exc()

finally:
    driver.quit()