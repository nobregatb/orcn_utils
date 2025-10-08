from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os
import time
import re
from datetime import datetime

CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
FILES_FOLDER = r"C:\Users\tbnobrega\OneDrive - ANATEL\Anatel\_ORCN"
PROFILE_DIR = os.path.abspath("./meu_perfil_chrome")
LOG_REQUERIMENTOS = os.path.join(FILES_FOLDER, 'ORCN.xlsx')


def criar_pasta_se_nao_existir(req):
    """Cria pasta do requerimento se não existir"""
    num, ano = req.split("/")
    requerimento = rf"Requerimentos\_{ano}.{num}"
    full_path = os.path.join(FILES_FOLDER, requerimento)
    if not os.path.exists(full_path):
        os.makedirs(full_path)
        print(f"   📁 Pasta criada: {full_path}")
    return full_path


def atualizar_excel(rows) -> list:
    """Atualiza planilha com novos requerimentos em análise"""
    novosAnalises = []
    wb = load_workbook(LOG_REQUERIMENTOS)
    ws = wb['Requerimentos-Análise']
    tabRequerimentos = ws.tables['tabRequerimentos']
    
    for i, row in enumerate(rows, start=1):
        cols = row.query_selector_all("td") 
        dados = [col.inner_text().strip() for col in cols]
        dados[0] = 'AUTOMATICO'
        
        linhas_existentes = [tuple(row) for row in ws.iter_rows(values_only=True)]
        
        if 1==1: #dados[9] == 'Em Análise':
            novalinha = dados[:10]
            if tuple(novalinha) not in linhas_existentes:
                ws.append(novalinha)
                print(f"   ✅ Linha adicionada: {novalinha[1]}")
                novosAnalises.append(novalinha[1])
                criar_pasta_se_nao_existir(novalinha[1])
            else:
                print(f"   ⏭️  Linha já existe: {novalinha[1]}")
    
    # Atualiza referência da tabela
    ref_atual = tabRequerimentos.ref
    ultima_linha = ws.max_row
    col_inicio, linha_inicio = ref_atual.split(":")[0][0], ref_atual.split(":")[0][1:]
    col_fim = ref_atual.split(":")[1][0]
    nova_ref = f"{col_inicio}{linha_inicio}:{col_fim}{ultima_linha}"
    tabRequerimentos.ref = nova_ref

    wb.save(LOG_REQUERIMENTOS)
    wb.close()
    
    print(f"\n📊 Total de novos requerimentos: {len(novosAnalises)}")
    return novosAnalises


def wait_primefaces_ajax(page, timeout=5000):
    """Espera todas as requisições AJAX do PrimeFaces terminarem"""
    try:
        page.wait_for_function(
            """() => {
                if (typeof PrimeFaces === 'undefined') return true;
                if (!PrimeFaces.ajax) return true;
                if (!PrimeFaces.ajax.Queue) return true;
                return PrimeFaces.ajax.Queue.isEmpty();
            }""",
            timeout=timeout
        )
    except:
        pass
    time.sleep(0.3)


def primefaces_click(page, element, description="elemento"):
    """
    Clica em botões submit do PrimeFaces (type="submit").
    Esses botões precisam submeter o formulário via AJAX.
    """
    print(f"   🎯 Clicando em: {description}")
    
    # Scroll até o elemento
    try:
        element.scroll_into_view_if_needed()
        time.sleep(0.3)
    except:
        pass
    
    # MÉTODO 1: Executa o onclick diretamente se existir
    try:
        onclick_executed = page.evaluate("""(button) => {
            const onclick = button.getAttribute('onclick');
            if (onclick) {
                try {
                    // Executa o código onclick
                    eval(onclick);
                    return true;
                } catch (e) {
                    console.error('Erro ao executar onclick:', e);
                    return false;
                }
            }
            return false;
        }""", element)
        
        if onclick_executed:
            print(f"   ✅ Onclick executado diretamente")
            time.sleep(1)
            return True
    except Exception as e:
        print(f"   ⚠ Onclick falhou: {str(e)[:50]}")
    
    # MÉTODO 2: Submit via PrimeFaces.ajax.Request
    try:
        success = page.evaluate("""(button) => {
            try {
                const form = button.closest('form');
                if (!form) return false;
                
                // Cria input hidden com dados do botão
                const hiddenInput = document.createElement('input');
                hiddenInput.type = 'hidden';
                hiddenInput.name = button.name || button.id;
                hiddenInput.value = button.value || button.id;
                form.appendChild(hiddenInput);
                
                // Usa PrimeFaces.ajax.Request se disponível
                if (typeof PrimeFaces !== 'undefined' && PrimeFaces.ajax && PrimeFaces.ajax.Request) {
                    const options = {
                        source: button.id,
                        process: button.id,
                        update: '@form',
                        formId: form.id,
                        params: []
                    };
                    
                    const paramObj = {};
                    paramObj[button.name || button.id] = button.value || button.id;
                    options.params.push(paramObj);
                    
                    PrimeFaces.ajax.Request.handle(options);
                    return true;
                }
                
                // Fallback: dispara submit
                const submitEvent = new Event('submit', {
                    bubbles: true,
                    cancelable: true
                });
                
                form._submitButton = button;
                form.dispatchEvent(submitEvent);
                
                if (!submitEvent.defaultPrevented) {
                    if (typeof jsf !== 'undefined' && jsf.ajax && jsf.ajax.request) {
                        jsf.ajax.request(button, null, {
                            'javax.faces.behavior.event': 'action'
                        });
                        return true;
                    }
                    form.submit();
                }
                
                return true;
            } catch (e) {
                console.error('Erro ao submeter:', e);
                return false;
            }
        }""", element)
        
        if success:
            print(f"   ✅ Submit executado")
            time.sleep(1)
            return True
        else:
            print(f"   ⚠ Submit falhou")
    except Exception as e:
        print(f"   ❌ Erro: {str(e)[:80]}")
    
    # MÉTODO 3: Force click como último recurso
    try:
        print(f"   🔄 Tentando force click...")
        element.click(force=True, timeout=2000)
        print(f"   ✅ Force click funcionou")
        time.sleep(1)
        return True
    except Exception as e:
        print(f"   ❌ Force click falhou: {str(e)[:50]}")
    
    return False


def baixar_pdfs(page, requerimento):
    """Baixa todos os PDFs da página de anexos"""
    num, ano = requerimento.split("/")
    pasta_destino = os.path.join(FILES_FOLDER, f"Requerimentos\\_{ano}.{num}")
    
    print(f"   📥 Buscando PDFs para baixar...")
    
    # Lista de botões que revelam PDFs
    botoes_pdf = [
        "Outros",
        "ART", 
        "Selo ANATEL",
        "Relatório de Avaliação da Conformidade - RACT",
        "Manual do Produto",
        "Certificado de Conformidade Técnica - CCT",
        "Contrato Social",
        "Fotos internas",
        "Relatório de Ensaio",
        "Fotos do produto"
    ]
    
    total_pdfs_baixados = 0
    
    # Percorre cada botão para revelar PDFs
    for nome_botao in botoes_pdf:
        print(f"   🔍 Procurando botão: {nome_botao}")
        
        try:
            # Busca o botão pelo nome/texto
            botao = page.get_by_role("button", name=nome_botao)
            
            # Verifica se o botão existe e está visível
            if botao.count() > 0:
                print(f"   🎯 Clicando em: {nome_botao}")
                
                # Clica no botão
                botao.first.click()
                
                # Aguarda o carregamento dos PDFs
                time.sleep(1)
                wait_primefaces_ajax(page)
                
                # Busca todos os links de PDF que foram revelados
                pdf_links = page.query_selector_all("a[href*='.pdf'], a[href*='download']")

                # Extrai informações da tabela
                tabela = page.query_selector("table.analiseTable")
                linhas_dados = []
                
                if tabela:
                    # Pega todas as linhas da tabela (exceto cabeçalho)
                    linhas = tabela.query_selector_all("tr")
                    
                    # Identifica o cabeçalho (primeira linha)
                    if len(linhas) > 0:
                        cabecalho = linhas[0].query_selector_all("th, td")
                        headers = [col.inner_text().strip() for col in cabecalho]
                        print(f"   📋 Cabeçalho da tabela: {headers}")
                        
                        # Processa as linhas de dados (exceto cabeçalho)
                        for linha in linhas[1:]:
                            colunas = linha.query_selector_all("th, td")
                            dados = [col.inner_text().strip() for col in colunas]
                            
                            if len(dados) >= len(headers):
                                linha_info = {}
                                for i, header in enumerate(headers):
                                    if i < len(dados):
                                        linha_info[header] = dados[i]
                                linhas_dados.append(linha_info)
                    
                    print(f"   📊 {len(linhas_dados)} linha(s) de dados extraída(s)")
                
                if pdf_links:
                    print(f"   📄 {len(pdf_links)} PDF(s) encontrado(s) para {nome_botao}")
                    
                    # Verifica se há correspondência entre PDFs e linhas da tabela
                    if len(pdf_links) != len(linhas_dados):
                        print(f"   ⚠ AVISO: {len(pdf_links)} PDFs mas {len(linhas_dados)} linhas na tabela!")
                    
                    # Baixa cada PDF encontrado
                    for idx, link in enumerate(pdf_links):
                        try:
                            # Verifica se o link está visível/disponível
                            if not link.is_visible():
                                continue
                            
                            # Pega informações da linha correspondente da tabela
                            linha_info = None
                            if idx < len(linhas_dados):
                                linha_info = linhas_dados[idx]
                            
                            # Pega o nome do arquivo original
                            href = link.get_attribute('href')
                            texto = link.inner_text().strip()
                            
                            # Configura download
                            with page.expect_download() as download_info:
                                link.click()
                            
                            download = download_info.value
                            
                            # Define nome do arquivo baseado na tabela
                            nome_arquivo = download.suggested_filename
                            if not nome_arquivo or nome_arquivo == "":
                                nome_arquivo = f"anexo_{idx + 1}.pdf"
                            
                            # Extrai nome base e extensão
                            nome_base, extensao = os.path.splitext(nome_arquivo)
                            
                            # Gera novo nome baseado nas informações da tabela
                            if linha_info:
                                try:
                                    # Extrai informações da tabela
                                    doc_id = linha_info.get("ID", f"#{idx + 1}")
                                    tipo_doc = linha_info.get("Tipo de Documento", "Documento")
                                    data_hora = linha_info.get("Data - Hora", "")
                                    
                                    # Processa a data para formato yyyy.mm.dd
                                    data_formatada = ""
                                    if data_hora:
                                        try:
                                            # Remove espaços extras e separa data da hora
                                            data_parte = data_hora.split()[0] if data_hora else ""
                                            
                                            # Padrões de data comuns
                                            padroes = [
                                                r"(\d{2})/(\d{2})/(\d{4})",  # dd/mm/yyyy
                                                r"(\d{4})-(\d{2})-(\d{2})",  # yyyy-mm-dd
                                                r"(\d{2})-(\d{2})-(\d{4})",  # dd-mm-yyyy
                                            ]
                                            
                                            for padrao in padroes:
                                                match = re.search(padrao, data_parte)
                                                if match:
                                                    if padrao == padroes[0] or padrao == padroes[2]:  # dd/mm/yyyy ou dd-mm-yyyy
                                                        dia, mes, ano = match.groups()
                                                    else:  # yyyy-mm-dd
                                                        ano, mes, dia = match.groups()
                                                    
                                                    data_formatada = f"{ano}.{mes.zfill(2)}.{dia.zfill(2)}"
                                                    break
                                            
                                            if not data_formatada:
                                                data_formatada = "0000.00.00"
                                                print(f"   ⚠ Não foi possível processar a data: {data_hora}")
                                        except Exception as e:
                                            data_formatada = "0000.00.00"
                                            print(f"   ⚠ Erro ao processar data: {str(e)[:50]}")
                                    else:
                                        data_formatada = "0000.00.00"
                                    
                                    # Limpa caracteres inválidos para nome de arquivo
                                    tipo_doc_limpo = re.sub(r'[<>:"/\\|?*]', '_', tipo_doc)
                                    doc_id_limpo = re.sub(r'[<>:"/\\|?*]', '_', str(doc_id))
                                    
                                    # Monta o novo nome: {yyyy.mm.dd}[{tipo} - ID {id}] {nome_base}{extensao}
                                    nome_arquivo = f"[{tipo_doc_limpo}][{data_formatada} - ID {doc_id_limpo}] {nome_base} [req {num} de  {ano}]{extensao}"
                                    
                                except Exception as e:
                                    print(f"   ⚠ Erro ao processar informações da tabela: {str(e)[:50]}")
                                    # Fallback para nome simples
                                    nome_arquivo = f"[{nome_botao}] {nome_base}{extensao}"
                            else:
                                # Fallback quando não há informação da tabela
                                nome_arquivo = f"[{nome_botao}] {nome_base}{extensao}"
                            
                            # Salva o arquivo
                            caminho_completo = os.path.join(pasta_destino, nome_arquivo)
                            download.save_as(caminho_completo)
                            
                            print(f"   ✅ Baixado: {nome_arquivo}")
                            total_pdfs_baixados += 1
                            
                        except Exception as e:
                            print(f"   ❌ Erro ao baixar PDF {idx + 1} de {nome_botao}: {str(e)[:50]}")
                else:
                    print(f"   ℹ Nenhum PDF encontrado para: {nome_botao}")
                    
            else:
                print(f"   ⚠ Botão não encontrado: {nome_botao}")
                
        except Exception as e:
            print(f"   ❌ Erro ao processar botão {nome_botao}: {str(e)[:50]}")
            continue
    
    if total_pdfs_baixados == 0:
        print(f"   ⚠ Nenhum PDF foi baixado")
    else:
        print(f"   💾 Total de {total_pdfs_baixados} PDF(s) salvos em: {pasta_destino}")


with sync_playwright() as p:
    browser = p.chromium.launch_persistent_context(
        PROFILE_DIR,
        headless=False,
        executable_path=CHROME_PATH,
        args=[
            "--start-maximized",
            "--disable-blink-features=AutomationControlled"
        ],
        accept_downloads=True  # IMPORTANTE: permite downloads
    )
    
    page = browser.new_page()
    
    # Navega para a lista
    lista_url = "https://sistemasnet.anatel.gov.br/mosaico/sch/worklist/"
    page.goto(lista_url)
    page.wait_for_load_state("load")
    
    # Clica em "Todos"
    page.click("#menuForm\\:todos",timeout=3600000) 
    page.wait_for_load_state("load")
    wait_primefaces_ajax(page)
    
    print("\n" + "="*60)
    print("🤖 AUTOMAÇÃO ORCN - DOWNLOAD DE ANEXOS")
    print("="*60)
    
    input("➡ Faça login (se necessário) e pressione ENTER para continuar...")
    
    # Seleciona todas as linhas
    rows = page.query_selector_all("css=#form\\:tarefasTable_data tr")
    print(f"\n🔎 {len(rows)} linhas encontradas na tabela")
    
    # Atualiza Excel e pega novos requerimentos
    print("\n📊 Atualizando planilha Excel...")
    novos_requerimentos = atualizar_excel(rows)
    
    if not novos_requerimentos:
        print("\n✅ Nenhum requerimento novo para processar!")
        input("Pressione ENTER para encerrar...")
        browser.close()
        exit()
    
    print(f"\n🚀 Processando {len(novos_requerimentos)} requerimento(s) novo(s)...\n")
    
    # Cria um dicionário com os dados de cada linha ANTES de iterar
    print("\n📋 Mapeando linhas da tabela...")
    linhas_dados = []
    
    for i, row in enumerate(rows, start=1):
        try:
            cols = row.query_selector_all("td")
            if len(cols) < 2:
                continue
            
            requerimento = cols[1].inner_text().strip()
            
            # Armazena os dados da linha
            linhas_dados.append({
                'indice': i,
                'requerimento': requerimento,
                'row': row
            })
        except Exception as e:
            print(f"   ⚠ Erro ao ler linha {i}: {str(e)[:50]}")
    
    print(f"   ✅ {len(linhas_dados)} linhas mapeadas")
    
    # Processa cada linha dos dados salvos
    for linha_info in linhas_dados:
        i = linha_info['indice']
        requerimento = linha_info['requerimento']
        
        # Só processa se for um requerimento novo
        if requerimento not in novos_requerimentos:
            continue
        
        print(f"\n{'='*60}")
        print(f"▶ LINHA {i}: {requerimento}")
        print(f"{'='*60}")
        
        # IMPORTANTE: Recarrega a linha atual usando um seletor mais específico
        # Busca pela linha que contém este requerimento específico
        try:
            row_atual = page.query_selector(f"css=#form\\:tarefasTable_data tr:has(td:text-is('{requerimento}'))")
            if not row_atual:
                # Fallback: busca todas as linhas novamente e pega pelo índice
                todas_linhas = page.query_selector_all("css=#form\\:tarefasTable_data tr")
                if i - 1 < len(todas_linhas):
                    row_atual = todas_linhas[i - 1]
                else:
                    print(f"   ⚠ Não foi possível encontrar a linha, pulando...")
                    continue
        except:
            print(f"   ⚠ Erro ao recarregar linha, pulando...")
            continue
        
        # Busca botão "Visualizar em Tela cheia" na linha atual 
        btn = row_atual.query_selector("button[type='submit'][title='Visualizar em Tela cheia']")
        if not btn:
            btn = row_atual.query_selector("button[title*='Tela cheia']")
        
        if not btn:
            print(f"   ⚠ Botão não encontrado, pulando...")
            continue
        
        # Clica no botão
        if not primefaces_click(page, btn, "Visualizar em Tela cheia"):
            print(f"   ⚠ Não foi possível clicar, pulando...")
            continue
        
        # Aguarda carregar
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except:
            pass
        wait_primefaces_ajax(page)
        
        print(f"   🌐 URL: {page.url}")
        
        # Busca botão "Anexos"
        time.sleep(2)
        print(f"   Buscando botão Anexos na toolbar...")
        
        iframe_element = page.wait_for_selector("#__frameDetalhe", timeout=10000)

        if iframe_element:
            detalhes_requerimento = iframe_element.get_attribute("src")
            #print(detalhes_requerimento)
            if detalhes_requerimento:
                page.goto(detalhes_requerimento)
                anexos_btn = page.get_by_role("button", name="Anexos")
                try:
                    #anexos_btn = page.query_selector("div.ui-toolbar-group-left button:has-text('Anexos')")
                    if anexos_btn:                        
                        anexos_btn.click(no_wait_after=True)
                        # Aguarda um seletor específico da nova “tela” ou da área que muda
                        page.wait_for_selector(".ui-blockui", state="detached", timeout=15000)
                        print("Página de anexos carregada.")
                    else:
                        print("Botão 'Anexos' não encontrado.")
                    wait_primefaces_ajax(page)
                        
                    print(f"   ✅ Página de Anexos carregada")
                        
                        # BAIXA OS PDFs
                    baixar_pdfs(page, requerimento)                        
                except Exception as e:
                    print(f"   ⚠ expect_navigation falhou: {str(e)}")
        else:
            print("⚠ iframe_element não encontrado, pulando...")
            continue

        #anexos_btn = None

        #primefaces_click
        #teogenes

        # Método 0: Botão com span contendo "Anexos"
        '''try:
            #anexos_btn = page.get_by_role("button", name="Anexos")
            #anexos_btn = page.query_selector("#formAnalise\\:j_idt110\\:2\\:j_idt115")
            #anexos_btn = page.query_selector("xpath=/html/body/span/form/span/div/div/div[1]/span/div/div/span[3]/span/button")
            #anexos_btn = page.query_selector("button[type='submit'][title='Anexos']")
            
            if anexos_btn:
                try:
                    btn_id = page.evaluate("(btn) => btn.id", anexos_btn)
                    print(f"   ID do botão: {btn_id}")
                except:
                    pass
                if primefaces_click(page, anexos_btn, "Anexos"):
                    try:
                        page.wait_for_load_state("networkidle", timeout=10000)
                    except:
                        pass
                    wait_primefaces_ajax(page)
                    
                    print(f"   ✅ Página de Anexos carregada")
                    
                    # BAIXA OS PDFs
                    baixar_pdfs(page, requerimento)
                else:
                    print(f"   ⚠ Não foi possível clicar em Anexos")
            else:
                print(f"   ⚠ Botão 'Anexos' não encontrado")                    
        except:
            pass'''


        
        
        
        # Volta para a lista
        print(f"   ↩ Voltando para a lista...")
        page.goto(lista_url, wait_until="networkidle")
        wait_primefaces_ajax(page)
        time.sleep(1)  # Pequena pausa para garantir que o DOM esteja estável
    
    print("\n" + "="*60)
    print("✅ PROCESSAMENTO CONCLUÍDO!")
    print("="*60)
    input("\nPressione ENTER para encerrar...")
    browser.close()